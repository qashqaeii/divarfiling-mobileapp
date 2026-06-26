import json
import os
import re
import time
import math
import sys
import threading
from typing import Any, Dict, List, Optional, Set

import requests


MAX_ITEMS_LIMIT = 500
MAX_LIST_PAGES = 60

_detail_request_lock = threading.Lock()
_detail_last_request_at = 0.0


# =========================
# پیکربندی‌های اولیه ساده
# =========================
CONFIG = {
    "city_ids": ["1"], 
    "category": "apartment-rent", 
    "district_ids": [],
    "sort": "sort_date",  
    "max_items": 100, 
    "sleep_between_requests_sec": 0.5,
    "timeout_sec": 20,
    "excel_output": "divar_items.xlsx",
    "max_workers": 6,
    "search_filters": {},
    "network_error": None,
    "last_request_error": None,
}


LIST_HEADERS = {
    "accept": "application/json, text/plain, */*",
    "accept-language": "en-US,en;q=0.9,fa;q=0.8",
    "content-type": "application/json",
    "origin": "https://divar.ir",
    "priority": "u=1, i",
    "referer": "https://divar.ir/",
    "sec-ch-ua": '"Google Chrome";v="141", "Not?A_Brand";v="8", "Chromium";v="141"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "same-site",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/141.0.0.0 Safari/537.36",
    "x-render-type": "CSR",
    "x-screen-size": "1536x247",
    "x-standard-divar-error": "true",
}

DETAIL_HEADERS = {
    "accept": "application/json-filled",
    "accept-language": "en-US,en;q=0.9,fa;q=0.8",
    "origin": "https://divar.ir",
    "priority": "u=1, i",
    "referer": "https://divar.ir/",
    "sec-ch-ua": '"Google Chrome";v="141", "Not?A_Brand";v="8", "Chromium";v="141"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "same-site",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/141.0.0.0 Safari/537.36",
    "x-render-type": "CSR",
    "x-screen-size": "1536x247",
}


def safe_get(d: Dict[str, Any], path: List[str], default: Any = None) -> Any:
    node: Any = d
    for key in path:
        if isinstance(node, dict) and key in node:
            node = node[key]
        else:
            return default
    return node


def to_english_digits(text: str) -> str:
    if not isinstance(text, str):
        return text
    # تبدیل اعداد فارسی/عربی به انگلیسی و حذف جداکننده‌های رایج
    persian = "۰۱۲۳۴۵۶۷۸۹"
    arabic = "٠١٢٣٤٥٦٧٨٩"
    trans_map = {ord(p): str(i) for i, p in enumerate(persian)}
    trans_map.update({ord(a): str(i) for i, a in enumerate(arabic)})
    trans_map.update({ord('٬'): ',', ord(','): ',', ord('\u200f'): None, ord('\u200e'): None})
    return text.translate(trans_map)


def parse_money_to_int(text: str) -> Optional[int]:
    if not text:
        return None
    t = to_english_digits(text)
    t = t.replace(' تومان', '').replace('تومان', '').replace(' ریال', '').replace('ریال', '')
    # حذف کاراکترهای غیر عددی بجز نقطه و علامت منفی
    digits = "-+.0123456789"
    cleaned = "".join(ch for ch in t if ch in digits)
    try:
        if cleaned == "":
            return None
        if "." in cleaned:
            return int(float(cleaned))
        return int(cleaned)
    except ValueError:
        return None


GEO_COORD_EXCEL_COLUMNS = ("عرض جغرافیایی", "طول جغرافیایی")
GEO_COORD_TEXT_FORMAT = "@"


def _normalize_header_text(text: Any) -> str:
    return (
        str(text or "")
        .strip()
        .replace("\u200c", "")
        .replace("\u200f", "")
        .replace("\ufeff", "")
    )


def _match_geo_header(header_text: Any) -> Optional[str]:
    normalized = _normalize_header_text(header_text)
    if not normalized:
        return None
    for geo_col in GEO_COORD_EXCEL_COLUMNS:
        if _normalize_header_text(geo_col) == normalized:
            return geo_col
    return None


def _format_geo_coord_text(val: Any) -> str:
    """مختصات را با دقت کامل به‌صورت متن برمی‌گرداند (مقاوم در برابر فرمت عدد صحیح اکسل)."""
    if val is None:
        return "-"
    try:
        import pandas as pd
        if pd.isna(val):
            return "-"
    except Exception:
        pass
    text = str(val).strip()
    if not text or text == "-":
        return "-"
    try:
        number = float(text.replace(",", "").replace(" ", ""))
    except (ValueError, TypeError):
        return text
    formatted = format(number, ".15f")
    return formatted.rstrip("0").rstrip(".") if "." in formatted else formatted


def _normalize_geo_coord(val: Any) -> str:
    return _format_geo_coord_text(val)


def _fix_worksheet_geo_coordinates(ws: Any, numeric_alignment: Any = None) -> None:
    """اجبار ذخیره مختصات به‌صورت متن — جلوگیری از رند شدن توسط فرمت #,##0."""
    from openpyxl.styles import Alignment

    align = numeric_alignment or Alignment(horizontal="center", vertical="center")
    geo_col_indexes = [
        col_idx
        for col_idx in range(1, ws.max_column + 1)
        if _match_geo_header(ws.cell(row=1, column=col_idx).value)
    ]
    if not geo_col_indexes:
        return
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in geo_col_indexes:
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None or str(cell.value).strip() in ("", "-"):
                continue
            cell.value = _format_geo_coord_text(cell.value)
            cell.number_format = GEO_COORD_TEXT_FORMAT
            cell.alignment = align


def _apply_geo_coord_format(cell: Any, header_text: str, numeric_alignment: Any) -> bool:
    """فرمت متنی برای ستون‌های عرض/طول جغرافیایی."""
    if not _match_geo_header(header_text):
        return False
    cell.alignment = numeric_alignment
    try:
        if cell.value is not None and str(cell.value).strip() not in ("", "-"):
            cell.value = _format_geo_coord_text(cell.value)
            cell.number_format = GEO_COORD_TEXT_FORMAT
    except Exception:
        pass
    return True


# ستون‌های شناخته‌شده ویژگی/امکانات برای خروجی اکسل (هر کدام در ستون جدا)
FEATURE_EXCEL_COLUMNS = [
    "تعداد واحد در طبقه",
    "تعداد کل طبقات ساختمان",
    "جهت ساختمان",
    "حیوان خانگی مجاز",
    "نوع آشپزخانه",
    "مناسب برای",
    "جنس کابینت",
    "بازسازی شده",
    "وضعیت واحد",
    "حداقل مدت قرارداد",
    "جنس کف",
    "بالکن",
    "سرویس بهداشتی",
    "سرمایش",
    "گرمایش",
    "تأمین‌کننده آب گرم",
    "تعداد حمام",
    "تعداد سرویس بهداشتی",
    "تعداد پارکینگ",
    "جکوزی",
    "سند",
    "سونا",
    "عرض ملک",
    "قابل معاوضه",
    "مجوز ساخت",
    "نمای ساختمان",
    "وضعیت سند",
    "وضعیت سکونت",
]

_FEATURE_ROW_PREFIXES = (
    "جنس کف ",
    "سرمایش ",
    "گرمایش ",
    "تأمین‌کننده آب گرم ",
    "سرویس بهداشتی ",
)


def _normalize_feature_row_title(title: str) -> tuple:
    """تبدیل عنوان FEATURE_ROW به (کلید ستون، مقدار)."""
    t = str(title or "").strip()
    if not t:
        return "", ""
    if "ودیعه" in t and "قابل تبدیل" in t:
        return "", ""
    if t.endswith(" ندارد"):
        return t[:-len(" ندارد")].strip(), "ندارد"
    if t.endswith(" دارد"):
        return t[:-len(" دارد")].strip(), "دارد"
    for prefix in _FEATURE_ROW_PREFIXES:
        if t.startswith(prefix):
            return prefix.strip(), t[len(prefix):].strip()
    return t, "دارد"


def _merge_feature_kv(target: Dict[str, str], source: Dict[str, str]) -> None:
    for key, val in source.items():
        k = str(key or "").strip()
        v = str(val or "").strip()
        if not k or not v:
            continue
        if k in target and target[k] != v:
            if v not in target[k]:
                target[k] = f"{target[k]} | {v}"
        else:
            target[k] = v


def _parse_group_feature_item(item: Dict[str, Any]) -> Dict[str, str]:
    title = str(item.get("title") or "").strip()
    if not title:
        return {}
    available = item.get("available", True)
    key, parsed_val = _normalize_feature_row_title(title)
    if not key:
        key = title
    if parsed_val in ("دارد", "ندارد"):
        val = parsed_val
    else:
        val = "دارد" if available else "ندارد"
    return {key: val}


def _parse_feature_widget_list(widget_list: List[Dict[str, Any]]) -> Dict[str, str]:
    """استخراج ویژگی‌ها از widget_list مودال یا صفحه جزئیات."""
    features: Dict[str, str] = {}
    pending_chip_label: Optional[str] = None

    for w in widget_list or []:
        wtype = w.get("widget_type")
        data = w.get("data") or {}

        if wtype == "UNEXPANDABLE_ROW":
            pending_chip_label = None
            title = str(data.get("title") or "").strip()
            value = str(data.get("value") or "").strip()
            if title and value:
                features[title] = value

        elif wtype == "FEATURE_ROW":
            pending_chip_label = None
            title = str(data.get("title") or "").strip()
            if not title:
                continue
            key, val = _normalize_feature_row_title(title)
            if not key:
                continue
            features[key] = val

        elif wtype == "DESCRIPTION_ROW":
            pending_chip_label = str(data.get("text") or "").strip() or None

        elif wtype == "WRAPPER_ROW":
            if pending_chip_label:
                chips = safe_get(data, ["chip_list", "chips"]) or []
                texts = [str(ch.get("text") or "").strip() for ch in chips if ch.get("text")]
                if texts:
                    features[pending_chip_label] = "، ".join(texts)
            pending_chip_label = None

        elif wtype in ("TITLE_ROW", "SECTION_TITLE_ROW"):
            pending_chip_label = None

    return features


def _extract_modal_features_from_action(action: Optional[Dict[str, Any]]) -> Dict[str, str]:
    if not action or action.get("type") != "LOAD_MODAL_PAGE":
        return {}
    modal_page = safe_get(action, ["payload", "modal_page"])
    if not modal_page:
        return {}
    return _parse_feature_widget_list(safe_get(modal_page, ["widget_list"]) or [])


def _default_business_lazy_body(token: str) -> Dict[str, Any]:
    return {
        "request_data": {
            "@type": "type.googleapis.com/premium_panel.GetPostBusinessLazyWidgetsRequest.RequestData",
            "post_token": token,
        }
    }


def _find_business_lazy_request(detail: Dict[str, Any]) -> Optional[tuple]:
    """مسیر و بدنهٔ درخواست lazy برای اطلاعات مشاور/آژانس از BUSINESS_SECTION."""
    for sec in detail.get("sections") or []:
        if sec.get("section_name") != "BUSINESS_SECTION":
            continue
        for w in sec.get("widgets") or []:
            if w.get("widget_type") != "LAZY_SECTION":
                continue
            data = w.get("data") or {}
            path = str(data.get("rest_request_path") or "").strip()
            if not path:
                continue
            req_data = data.get("request_data") or {}
            if req_data:
                body: Dict[str, Any] = {"request_data": req_data}
            else:
                post_token = str(safe_get(detail, ["webengage", "token"]) or "").strip()
                body = _default_business_lazy_body(post_token)
            return path, body
    return None


def fetch_business_lazy_widgets(
    detail: Dict[str, Any],
    token: str,
    session: Optional[requests.Session] = None,
) -> List[Dict[str, Any]]:
    """دریافت widget_list مشاور/آژانس از API lazy (POST)."""
    found = _find_business_lazy_request(detail)
    if found:
        path, body = found
    else:
        business_type = str(safe_get(detail, ["webengage", "business_type"]) or "").strip().lower()
        if business_type != "premium-panel":
            return []
        path = f"/v8/premium-user/post-page/business-data/{token}/lazy"
        body = _default_business_lazy_body(token)

    if not body.get("request_data"):
        body = _default_business_lazy_body(token)

    url = path if path.startswith("http") else f"https://api.divar.ir{path}"
    resp = request_with_retry(
        "POST",
        url,
        headers=LIST_HEADERS,
        json_body=body,
        timeout=CONFIG["timeout_sec"],
        session=session,
    )
    if resp is None or resp.status_code != 200:
        return []
    try:
        payload = resp.json()
    except json.JSONDecodeError:
        return []
    return payload.get("widget_list") or []


def _extract_consultant_from_widget_list(widget_list: List[Dict[str, Any]]) -> Dict[str, str]:
    """استخراج نام مشاور/آژانس از EVENT_ROW با PREMIUM_PANEL_BRAND_LANDING."""
    result = {
        "مشاور_نام": "",
        "مشاور_آگهی_فعال": "",
    }
    for w in widget_list or []:
        if w.get("widget_type") != "EVENT_ROW":
            continue
        data = w.get("data") or {}
        action = data.get("action") or {}
        if action.get("type") != "PREMIUM_PANEL_BRAND_LANDING":
            continue
        result["مشاور_نام"] = str(data.get("title") or "").strip()
        result["مشاور_آگهی_فعال"] = str(data.get("subtitle") or "").strip()
        break
    return result


def _apply_features_to_out(out: Dict[str, Any], features_kv: Dict[str, str]) -> List[str]:
    """اعمال features_kv روی out و برگرداندن لیست عنوان‌ها برای ستون تجمیعی."""
    titles: List[str] = []
    for key, val in features_kv.items():
        out[key] = val
        titles.append(f"{key}: {val}" if val not in ("دارد", "ندارد") else f"{key} {val}")

    for amenity in ("آسانسور", "پارکینگ", "انباری"):
        if amenity in features_kv and features_kv[amenity] in ("دارد", "ندارد"):
            out[amenity] = features_kv[amenity]

    return titles


def _discover_extra_feature_columns(df: Any) -> List[str]:
    """ستون‌های ویژگی اضافی که در داده‌ها هست ولی در لیست ثابت نیست."""
    if df is None or not len(getattr(df, "columns", [])):
        return []
    skip_exact = {
        "token", "title", "description", "images", "latitude", "longitude",
        "address", "share_link", "business_type", "status", "phone_text",
        "category_slug", "لینک", "شناسه", "عنوان", "توضیحات", "زیرعنوان",
        "برچسب‌ها", "attributes", "ویژگی_ها_و_امکانات", "created_at",
        "city_persian", "district_persian", "normalized_city", "normalized_district",
        "price_text", "آسانسور", "پارکینگ", "انباری",
    }
    skip_prefixes = (
        "token", "title", "description", "category", "price", "images",
        "latitude", "longitude", "address", "city_", "district", "share",
        "business", "status", "phone", "مالی_", "اجاره_", "مکان_", "نقشه_",
        "رسانه_", "فروشنده_", "مشاور_", "سیستمی_", "انتشار_", "تماس_", "دسته‌بندی",
        "برچسب", "created", "normalized_", "category_",
    )
    known = set(FEATURE_EXCEL_COLUMNS)
    extras: List[str] = []
    for col in df.columns:
        c = str(col)
        if c in known or c in skip_exact:
            continue
        if any(c.startswith(p) for p in skip_prefixes):
            continue
        if c.startswith("ویژگی"):
            continue
        extras.append(c)
    return sorted(set(extras))


def _append_feature_detail_columns(result_df: Any, df: Any, combine_excel_and_source) -> List[str]:
    """افزودن ستون‌های جداگانه ویژگی/امکانات به DataFrame خروجی."""
    added: List[str] = []
    ordered = list(FEATURE_EXCEL_COLUMNS) + _discover_extra_feature_columns(df)
    for col_name in ordered:
        if col_name not in df.columns and col_name not in getattr(result_df, "columns", []):
            continue
        if col_name in df.columns:
            result_df[col_name] = df[col_name]
        elif col_name in getattr(result_df, "columns", []):
            continue
        else:
            result_df[col_name] = combine_excel_and_source(col_name, col_name)
        if col_name not in added:
            added.append(col_name)
    return added


def _apply_search_filters(form_data: Dict[str, Any]) -> None:
    """اعمال فیلترهای پیشرفته از CONFIG['search_filters'] روی form_data دیوار."""
    sf = CONFIG.get("search_filters") or {}
    if not sf:
        return

    def _num_range(key: str, min_key: str, max_key: str) -> None:
        lo = sf.get(min_key)
        hi = sf.get(max_key)
        if lo is None and hi is None:
            return
        rng: Dict[str, Any] = {}
        if lo is not None:
            rng["minimum"] = int(lo)
        if hi is not None:
            rng["maximum"] = int(hi)
        if rng:
            form_data[key] = {"number_range": rng}

    cat = str(CONFIG.get("category") or "")
    is_rent = "rent" in cat
    if is_rent:
        _num_range("credit", "deposit_min", "deposit_max")
        _num_range("rent", "rent_min", "rent_max")
    else:
        _num_range("price", "price_min", "price_max")

    _num_range("size", "area_min", "area_max")
    _num_range("building-year", "year_min", "year_max")

    rooms = sf.get("rooms")
    if rooms:
        if isinstance(rooms, str):
            rooms = [r.strip() for r in rooms.split(",") if r.strip()]
        if rooms:
            form_data["rooms"] = {"repeated_string": {"value": [str(r) for r in rooms]}}


def build_search_payload(page: int, last_post_date: Optional[str], search_uid: Optional[str], cumulative_widgets_count: int, viewed_tokens: Optional[str] = None) -> Dict[str, Any]:
    form_data: Dict[str, Any] = {
        "category": {"str": {"value": CONFIG["category"]}},
    }

    if CONFIG["district_ids"]:
        form_data["districts"] = {"repeated_string": {"value": CONFIG["district_ids"]}}

    _apply_search_filters(form_data)

    pagination_data: Dict[str, Any] = {
        "@type": "type.googleapis.com/post_list.PaginationData",
        "page": page,
        "layer_page": page,
        "cumulative_widgets_count": cumulative_widgets_count,
    }
    
    if last_post_date:
        pagination_data["last_post_date"] = last_post_date
    if search_uid:
        pagination_data["search_uid"] = search_uid
    if viewed_tokens:
        pagination_data["viewed_tokens"] = viewed_tokens

    payload: Dict[str, Any] = {
        "city_ids": CONFIG["city_ids"],
        "pagination_data": pagination_data,
        "disable_recommendation": False,
        "map_state": {"camera_info": {"bbox": {}}},
        "search_data": {
            "form_data": {"data": form_data},
            "server_payload": {
                "@type": "type.googleapis.com/widgets.SearchData.ServerPayload",
                "additional_form_data": {
                    "data": {"sort": {"str": {"value": CONFIG["sort"]}}}
                },
            },
        },
    }

    # سازگاری با برخی دسته‌ها که نیاز به source_view/page_state دارند
    commercial_categories = ("office-sell", "shop-sell", "industry-agriculture-business-sell")
    if CONFIG.get("category") in ("house-villa-sell", "house-villa-rent") or CONFIG.get("category") in commercial_categories:
        payload["source_view"] = "CATEGORY"
        try:
            payload["map_state"]["page_state"] = "HALF_STATE"
        except Exception:
            payload["map_state"] = {"camera_info": {"bbox": {}}, "page_state": "HALF_STATE"}

    return payload


def get_http_session() -> requests.Session:
    session = requests.Session()
    try:
        from requests.adapters import HTTPAdapter
        from urllib3.util.retry import Retry
        retries = Retry(total=2, backoff_factor=0.2, status_forcelist=[429, 500, 502, 503, 504])
        adapter = HTTPAdapter(pool_connections=CONFIG.get("max_workers", 6) * 2, pool_maxsize=CONFIG.get("max_workers", 6) * 2, max_retries=retries)
        session.mount("http://", adapter)
        session.mount("https://", adapter)
    except Exception:
        pass
    return session


def _set_last_request_error(**fields: Any) -> None:
    CONFIG["last_request_error"] = fields


def _extract_api_error_message(resp: requests.Response) -> str:
    try:
        body = resp.json()
    except (json.JSONDecodeError, ValueError):
        text = (resp.text or "").strip()
        return text[:200] if text else ""
    if isinstance(body, dict):
        msg = body.get("message")
        if isinstance(msg, dict):
            title = msg.get("title") or msg.get("message")
            if title:
                return str(title)
        if isinstance(msg, str) and msg:
            return msg
        err = body.get("error") or body.get("detail")
        if err:
            return str(err)
    return ""


def request_with_retry(method: str, url: str, *, headers: Dict[str, str], json_body: Optional[Dict[str, Any]] = None, timeout: int = 20, retries: int = 3, backoff_base: float = 0.8, session: Optional[requests.Session] = None) -> Optional[requests.Response]:
    last_exc: Optional[BaseException] = None
    for attempt in range(retries):
        try:
            if method.upper() == "POST":
                if session is not None:
                    resp = session.post(url, headers=headers, json=json_body, timeout=timeout)
                else:
                    resp = requests.post(url, headers=headers, json=json_body, timeout=timeout)
            else:
                if session is not None:
                    resp = session.get(url, headers=headers, timeout=timeout)
                else:
                    resp = requests.get(url, headers=headers, timeout=timeout)
            return resp
        except requests.exceptions.RequestException as exc:
            last_exc = exc
            _set_last_request_error(
                url=url,
                method=method.upper(),
                attempt=attempt + 1,
                retries=retries,
                timeout=timeout,
                error_type=type(exc).__name__,
                error=str(exc),
            )
            sleep_s = backoff_base * math.pow(2, attempt)
            time.sleep(sleep_s)
    if last_exc is not None:
        _set_last_request_error(
            url=url,
            method=method.upper(),
            attempt=retries,
            retries=retries,
            timeout=timeout,
            error_type=type(last_exc).__name__,
            error=str(last_exc),
        )
    return None


def fetch_post_tokens() -> List[str]:
    CONFIG["network_error"] = None
    CONFIG["last_request_error"] = None
    resume = CONFIG.get("resume_tokens")
    if resume:
        return list(resume)

    tokens: List[str] = []
    search_uid: Optional[str] = None
    last_post_date: Optional[str] = None
    cumulative_widgets_count: int = 0
    viewed_tokens: Optional[str] = None
    max_items = CONFIG.get("max_items", 100)
    max_items = max(1, min(MAX_ITEMS_LIMIT, int(max_items)))
    
    page = 1
    consecutive_empty_pages = 0  # تعداد صفحات متوالی خالی
    
    while len(tokens) < max_items:
        payload = build_search_payload(page, last_post_date, search_uid, cumulative_widgets_count, viewed_tokens)
        resp = request_with_retry(
            "POST",
            "https://api.divar.ir/v8/postlist/w/search",
            headers=LIST_HEADERS,
            json_body=payload,
            timeout=CONFIG["timeout_sec"],
        )

        time.sleep(CONFIG["sleep_between_requests_sec"])  # احترام به سرور

        if resp is None:
            detail = CONFIG.get("last_request_error") or {}
            err_type = detail.get("error_type", "RequestException")
            err_msg = detail.get("error", "")
            CONFIG["network_error"] = (
                "اتصال به سرور دیوار برقرار نشد. اتصال اینترنت را بررسی کنید."
                + (f" ({err_type}: {err_msg})" if err_msg else "")
            )
            print(
                f"[DIVAR] NETWORK FAIL url={detail.get('url', '')} "
                f"attempts={detail.get('retries', 3)} "
                f"timeout={detail.get('timeout', CONFIG['timeout_sec'])}s "
                f"error={err_type}: {err_msg}",
                file=sys.stderr,
            )
            break

        if resp.status_code != 200:
            api_msg = _extract_api_error_message(resp)
            category = CONFIG.get("category", "")
            CONFIG["network_error"] = (
                f"خطای API دیوار (کد {resp.status_code})"
                + (f": {api_msg}" if api_msg else "")
                + "."
            )
            _set_last_request_error(
                url="https://api.divar.ir/v8/postlist/w/search",
                method="POST",
                status_code=resp.status_code,
                category=category,
                api_message=api_msg,
                body=(resp.text or "")[:300],
            )
            print(
                f"[DIVAR] API ERROR status={resp.status_code} category={category} "
                f"message={api_msg or (resp.text or '')[:200]}",
                file=sys.stderr,
            )
            break

        try:
            data = resp.json()
        except json.JSONDecodeError:
            break


        if page == 1:
            map_data = data.get("map_data", {}) or {}
            post_count = map_data.get("post_count")
            if post_count is not None:
                CONFIG["total_posts_count"] = post_count

        list_widgets = data.get("list_widgets", []) or []
        page_tokens = []
        for w in list_widgets:
            # فقط POST_ROW را پردازش می‌کنیم (برای حفظ ترتیب دقیق)
            widget_type = w.get("widget_type", "")
            if widget_type == "POST_ROW":
                token = safe_get(w, ["data", "token"]) or safe_get(w, ["data", "action", "payload", "token"])
                if token and isinstance(token, str):
                    page_tokens.append(token)
        
        # حذف توکن‌های تکراری قبل از اضافه کردن
        existing_tokens_set = set(tokens)
        new_tokens = [t for t in page_tokens if t not in existing_tokens_set]
        
        # اضافه کردن توکن‌های جدید این صفحه (تا حد max_items)
        remaining = max_items - len(tokens)
        tokens.extend(new_tokens[:remaining])

        # به‌روزرسانی پارامترهای صفحه‌بندی
        if list_widgets:
            first = list_widgets[0]
            last = list_widgets[-1]
            info = safe_get(first, ["action_log", "server_side_info", "info"], {}) or {}
            # ISO تاریخ از آخرین ویجت برای صفحه بعدی
            last_info = safe_get(last, ["action_log", "server_side_info", "info"], {}) or {}
            last_post_date = last_info.get("sort_date") or info.get("sort_date") or last_post_date
            extra = info.get("extra_data", {}) or {}
            search_uid = extra.get("search_uid") or search_uid
            # cumulative_widgets_count: مجموع تعداد ویجت‌های تمام صفحات قبلی + صفحه فعلی
            cumulative_widgets_count += len(list_widgets)
            # viewed_tokens از pagination_data پاسخ (اگر موجود باشد)
            pagination = data.get("pagination_data", {})
            if pagination and "viewed_tokens" in pagination:
                viewed_tokens = pagination.get("viewed_tokens")
            consecutive_empty_pages = 0  # ریست کردن شمارنده صفحات خالی
        else:
            consecutive_empty_pages += 1
            # اگر دو صفحه متوالی خالی بودند، متوقف شویم
            if consecutive_empty_pages >= 2:
                break

        # اگر به تعداد مورد نظر رسیدیم، متوقف شویم
        if len(tokens) >= max_items:
            break
        
        page += 1
        if page > MAX_LIST_PAGES:
            break

    # حذف توکن‌های تکراری و محدود کردن به max_items
    unique_tokens = list(dict.fromkeys(tokens))[:max_items]
    return unique_tokens


def fetch_post_detail(token: str, session: Optional[requests.Session] = None) -> Optional[Dict[str, Any]]:
    global _detail_last_request_at
    delay = float(CONFIG.get("sleep_between_requests_sec", 0.5))
    with _detail_request_lock:
        wait = delay - (time.monotonic() - _detail_last_request_at)
        if wait > 0:
            time.sleep(wait)
        _detail_last_request_at = time.monotonic()

    url = f"https://api.divar.ir/v8/posts-v2/web/{token}"
    resp = request_with_retry(
        "GET",
        url,
        headers=DETAIL_HEADERS,
        timeout=CONFIG["timeout_sec"],
        session=session,
    )
    if resp is None or resp.status_code != 200:
        return None
    try:
        detail = resp.json()
    except json.JSONDecodeError:
        return None

    token_from_detail = safe_get(detail, ["webengage", "token"]) or token
    try:
        widget_list = fetch_business_lazy_widgets(detail, token_from_detail, session)
        if widget_list:
            detail["business_lazy_widget_list"] = widget_list
    except Exception:
        pass
    return detail


def flatten_post_detail(detail: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}

    seo = detail.get("seo", {}) or {}
    share = detail.get("share", {}) or {}
    webengage = detail.get("webengage", {}) or {}
    analytics = detail.get("analytics", {}) or {}
    sections = detail.get("sections", []) or []

    # شناسه و لینک
    token_val = webengage.get("token") or (safe_get(share, ["web_url"]).split("/")[-1] if share.get("web_url") else "")
    out["token"] = token_val
    out["شناسه"] = token_val
    out["لینک"] = safe_get(share, ["web_url"]) or (f"https://divar.ir/v/{token_val}" if token_val else "")

    # عنوان و زیرعنوان
    title_from_sections = ""
    subtitle_text = ""
    for s in sections:
        if s.get("section_name") == "TITLE":
            for w in s.get("widgets", []) or []:
                if w.get("widget_type") in ("LEGEND_TITLE_ROW", "TITLE_ROW"):
                    title_from_sections = safe_get(w, ["data", "title"]) or safe_get(w, ["data", "text"]) or title_from_sections
                    if w.get("widget_type") == "LEGEND_TITLE_ROW":
                        subtitle_text = safe_get(w, ["data", "subtitle"]) or subtitle_text
            break
    final_title = title_from_sections or safe_get(seo, ["web_info", "title"]) or safe_get(seo, ["title"]) or ""
    out["title"] = final_title
    out["عنوان"] = final_title
    out["زیرعنوان"] = subtitle_text

    # توضیحات
    description_text = ""
    for s in sections:
        if s.get("section_name") == "DESCRIPTION":
            for w in s.get("widgets", []) or []:
                if w.get("widget_type") in ("DESCRIPTION_ROW",):
                    description_text = safe_get(w, ["data", "text"]) or ""
                    if description_text:
                        break
            break
    out["description"] = description_text
    out["توضیحات"] = description_text

    # شهر/محله و دسته‌بندی
    city_pers = safe_get(seo, ["web_info", "city_persian"]) or ""
    dist_pers = safe_get(seo, ["web_info", "district_persian"]) or ""
    out["city_persian"] = city_pers
    out["district_persian"] = dist_pers
    out["normalized_city"] = analytics.get("city") or ""
    out["normalized_district"] = webengage.get("district") or ""
    out["مکان_شهر"] = city_pers
    out["مکان_منطقه"] = dist_pers

    cat1 = analytics.get("cat1") or ""
    cat2 = analytics.get("cat2") or ""
    cat3 = analytics.get("cat3") or webengage.get("category") or ""
    out["category_slug"] = cat3 or cat2 or cat1
    out["دسته‌بندی_سطح۱"] = cat1
    out["دسته‌بندی_سطح۲"] = cat2
    out["دسته‌بندی_سطح۳"] = cat3
    out["دسته‌بندی_عنوان"] = safe_get(seo, ["web_info", "category_slug_persian"]) or ""

    # قیمت‌ها و مشخصات از LIST_DATA
    total_price = ""
    price_per_meter = ""
    total_price_num: Optional[int] = None
    price_per_meter_num: Optional[int] = None
    # فیلدهای مخصوص اجاره
    deposit_text = ""
    rent_monthly_text = ""
    deposit_num: Optional[int] = None
    rent_monthly_num: Optional[int] = None
    rent_convertibility = ""  # قابل تبدیل/غیر قابل تبدیل
    attrs_list: List[str] = []
    # مقادیر کمکی برای اسلایدر اجاره (در برخی آگهی‌ها ردیف‌های «ودیعه/اجاره» صریح نیست)
    slider_credit_value: Optional[int] = None
    slider_rent_value: Optional[int] = None
    slider_credit_transformed: Optional[int] = None
    slider_rent_transformed: Optional[int] = None
    has_rent_slider = False  # فلگ برای تشخیص وجود RENT_SLIDER
    for s in sections:
        if s.get("section_name") == "LIST_DATA":
            for w in s.get("widgets", []) or []:
                wtype = w.get("widget_type")
                if wtype == "GROUP_INFO_ROW":
                    items = safe_get(w, ["data", "items"]) or []
                    if isinstance(items, list):
                        for it in items:
                            t = str(it.get("title") or "").strip()
                            v = str(it.get("value") or "").strip()
                            if t or v:
                                attrs_list.append(f"{t}:{v}")
                elif wtype == "UNEXPANDABLE_ROW":
                    title = str(safe_get(w, ["data", "title"]) or "").strip()
                    value = str(safe_get(w, ["data", "value"]) or "").strip()
                    if title == "قیمت کل":
                        total_price = value
                        total_price_num = parse_money_to_int(value)
                    elif title == "قیمت هر متر":
                        price_per_meter = value
                        price_per_meter_num = parse_money_to_int(value)
                    elif title in ("ودیعه", "بیعانه", "رهن", "رهن کامل"):
                        deposit_text = value
                        deposit_num = parse_money_to_int(value)
                    elif title in ("اجارهٔ ماهانه", "اجاره ماهانه"):
                        rent_monthly_text = value
                        rent_monthly_num = parse_money_to_int(value)
                    elif title in ("ودیعه و اجاره", "رهن و اجاره"):
                        rent_convertibility = value
                    elif title or value:
                        attrs_list.append(f"{title}:{value}")
                elif wtype == "FEATURE_ROW":
                    # برخی آگهی‌ها فقط یک متن دربارهٔ قابلیت تبدیل دارند
                    # فقط اگر rent_convertibility قبلاً تنظیم نشده باشد
                    if not rent_convertibility:
                        fr_title = str(safe_get(w, ["data", "title"]) or "").strip()
                        if fr_title:
                            # بررسی "قابل تبدیل"
                            if "قابل تبدیل" in fr_title:
                                rent_convertibility = "قابل تبدیل"
                            # بررسی "غیر قابل تبدیل" (برای اطمینان)
                            elif "غیر قابل تبدیل" in fr_title or "غیرقابل تبدیل" in fr_title:
                                rent_convertibility = "غیر قابل تبدیل"
                elif wtype == "RENT_SLIDER":
                    # RENT_SLIDER نشان‌دهنده قابل تبدیل بودن است
                    has_rent_slider = True
                    c_val = safe_get(w, ["data", "credit", "value"]) or ""
                    c_tr = safe_get(w, ["data", "credit", "transformed_value"]) or ""
                    r_val = safe_get(w, ["data", "rent", "value"]) or ""
                    r_tr = safe_get(w, ["data", "rent", "transformed_value"]) or ""
                    slider_credit_value = parse_money_to_int(str(c_val))
                    slider_credit_transformed = parse_money_to_int(str(c_tr))
                    slider_rent_value = parse_money_to_int(str(r_val))
                    slider_rent_transformed = parse_money_to_int(str(r_tr))
    out["price_text"] = total_price
    out["مالی_قیمت_کل_نمایش"] = total_price
    out["مالی_قیمت_کل_عدد"] = total_price_num if total_price_num is not None else webengage.get("price")
    if price_per_meter:
        attrs_list.append(f"قیمت هر متر:{price_per_meter}")
    out["مالی_قیمت_هر_متر_نمایش"] = price_per_meter
    out["مالی_قیمت_هر_متر_عدد"] = price_per_meter_num
    out["مالی_ارز"] = "تومان" if (total_price or price_per_meter) else ""
    out["مالی_منبع_قیمت_وب‌انگیج"] = webengage.get("price")

    # اجاره
    # تکمیل از RENT_SLIDER در صورت نبود دادهٔ صریح
    if deposit_num is None:
        # ابتدا transformed سپس value
        deposit_num = slider_credit_transformed if slider_credit_transformed is not None else slider_credit_value
    if rent_monthly_num is None:
        rent_monthly_num = slider_rent_transformed if slider_rent_transformed is not None else slider_rent_value
    # تکمیل از webengage در صورت نبود
    if deposit_num is None:
        deposit_num = webengage.get("credit") if isinstance(webengage.get("credit"), int) else None
    if rent_monthly_num is None:
        rent_monthly_num = webengage.get("rent") if isinstance(webengage.get("rent"), int) else None

    out["اجاره_ودیعه_نمایش"] = deposit_text
    out["اجاره_ودیعه_عدد"] = deposit_num
    out["اجاره_ماهانه_نمایش"] = rent_monthly_text
    out["اجاره_ماهانه_عدد"] = rent_monthly_num

    if rent_convertibility:
        out["اجاره_قابل_تبدیل"] = rent_convertibility
    elif has_rent_slider:
        out["اجاره_قابل_تبدیل"] = "قابل تبدیل"
    else:
        out["اجاره_قابل_تبدیل"] = ""

    # ویژگی‌ها و امکانات (صفحه اصلی + مودال LOAD_MODAL_PAGE)
    features_kv: Dict[str, str] = {}
    features_titles: List[str] = []
    for s in sections:
        if s.get("section_name") != "LIST_DATA":
            continue
        for w in s.get("widgets", []) or []:
            wtype = w.get("widget_type")
            data = w.get("data") or {}

            if wtype == "GROUP_FEATURE_ROW":
                for it in data.get("items") or []:
                    _merge_feature_kv(features_kv, _parse_group_feature_item(it))
                _merge_feature_kv(features_kv, _extract_modal_features_from_action(data.get("action")))

            elif wtype == "SELECTOR_ROW":
                selector_title = str(data.get("title") or "").strip()
                modal_features = _extract_modal_features_from_action(data.get("action"))
                if modal_features:
                    _merge_feature_kv(features_kv, modal_features)
                elif selector_title and "ویژگی" in selector_title:
                    pass  # مودال lazy — در پاسخ detail معمولاً embed است

            elif wtype == "FEATURE_ROW":
                title = str(data.get("title") or "").strip()
                if title and "قابل تبدیل" not in title:
                    key, val = _normalize_feature_row_title(title)
                    if key:
                        features_kv[key] = val
                        attrs_list.append(f"امکانات:{title}")

    features_titles = _apply_features_to_out(out, features_kv)

    # مختصات و نقشه
    lat = ""
    lon = ""
    map_type = ""
    map_radius = ""
    map_image = ""
    for s in sections:
        if s.get("section_name") == "MAP":
            for w in s.get("widgets", []) or []:
                if w.get("widget_type") == "MAP_ROW":
                    lat = str(
                        safe_get(w, ["data", "location", "exact_data", "point", "latitude"]) or
                        safe_get(w, ["data", "location", "fuzzy_data", "point", "latitude"]) or ""
                    )
                    lon = str(
                        safe_get(w, ["data", "location", "exact_data", "point", "longitude"]) or
                        safe_get(w, ["data", "location", "fuzzy_data", "point", "longitude"]) or ""
                    )
                    map_type = safe_get(w, ["data", "location", "type"]) or ""
                    map_radius = str(safe_get(w, ["data", "location", "fuzzy_data", "radius"]) or "")
                    map_image = safe_get(w, ["data", "image_url"]) or ""
                    break
    out["latitude"] = lat
    out["longitude"] = lon
    out["مکان_مختصات_lat"] = lat
    out["مکان_مختصات_lng"] = lon
    out["نقشه_نوع"] = map_type
    out["نقشه_شعاع_محدوده_متر"] = map_radius
    out["نقشه_تصویر"] = map_image

    # تصاویر و ویدیو
    image_urls: List[str] = []
    video_url = ""
    for s in sections:
        if s.get("section_name") == "IMAGE":
            for w in s.get("widgets", []) or []:
                if w.get("widget_type") == "IMAGE_CAROUSEL":
                    items = safe_get(w, ["data", "items"]) or []
                    for it in items or []:
                        url = safe_get(it, ["image", "url"]) or ""
                        if url:
                            image_urls.append(url)
                        vurl = safe_get(it, ["video_url"]) or ""
                        if vurl:
                            video_url = vurl
            break
    out["images"] = " | ".join(image_urls)
    out["رسانه_تصاویر"] = " | ".join(image_urls)
    out["رسانه_تعداد_تصویر"] = len(image_urls) if image_urls else ""
    out["رسانه_ویدیو"] = video_url
    # اصالت تصویر
    image_originality = ""
    for s in sections:
        if s.get("section_name") == "LIST_DATA":
            for w in s.get("widgets", []) or []:
                if w.get("widget_type") == "UNEXPANDABLE_ROW":
                    title = str(safe_get(w, ["data", "title"]) or "")
                    if title.strip() == "تصویر‌ها برای همین ملک است؟":
                        image_originality = str(safe_get(w, ["data", "value"]) or "")
                        break
    out["رسانه_اصالت_تصویر_برای_همین_ملک"] = image_originality

    # لینک اشتراک
    out["share_link"] = safe_get(share, ["web_url"]) or ""

    # فروشنده
    out["business_type"] = webengage.get("business_type") or ""
    out["status"] = webengage.get("status") or ""
    out["فروشنده_business_type"] = webengage.get("business_type") or ""
    out["فروشنده_نوع"] = ""
    out["فروشنده_تخفیف"] = "جزئی" if "تخفیف" in (description_text or "") else ""

    # مشاور / آژانس املاک (از API lazy: business-data/.../lazy)
    business_widgets = detail.get("business_lazy_widget_list") or detail.get("widget_list") or []
    consultant = _extract_consultant_from_widget_list(business_widgets)
    out.update(consultant)
    out["مشاور_business_ref"] = webengage.get("business_ref") or ""

    # آدرس (از توضیحات اگر یافت شود)
    address = "، ".join([p for p in [city_pers, dist_pers] if p])
    if description_text and "آدرس" in description_text:
        for line in description_text.splitlines():
            if line.strip().startswith("آدرس"):
                addr = line.split(":", 1)
                if len(addr) == 2:
                    address = addr[1].strip()
                break
    out["address"] = address
    out["مکان_آدرس"] = address

    # برچسب‌ها
    tags: List[str] = []
    for s in sections:
        if s.get("section_name") == "TAGS":
            for w in s.get("widgets", []) or []:
                if w.get("widget_type") == "WRAPPER_ROW":
                    chips = safe_get(w, ["data", "chip_list", "chips"]) or []
                    for ch in chips or []:
                        t = ch.get("text")
                        if t:
                            tags.append(t)
    out["برچسب‌ها"] = " | ".join(tags)

    # انتشار
    out["created_at"] = ""
    out["انتشار_تاریخ_شمسی"] = ""
    out["انتشار_تاریخ_میلادی"] = ""
    out["انتشار_unavailable_after"] = safe_get(seo, ["unavailable_after"]) or ""

    # تماس
    contact = detail.get("contact", {}) or {}
    out["phone_text"] = ""
    out["تماس_چت_فعال"] = contact.get("chat_enabled")
    out["تماس_تماس_ایمن"] = contact.get("secure_call_enabled")
    out["تماس_contact_uuid"] = contact.get("contact_uuid")

    # سیستمی
    out["سیستمی_webengage_category"] = webengage.get("category") or ""
    out["سیستمی_webengage_city"] = webengage.get("city") or ""
    out["سیستمی_webengage_district"] = webengage.get("district") or ""
    out["سیستمی_webengage_price"] = webengage.get("price")
    out["سیستمی_webengage_image_count"] = webengage.get("image_count")
    out["سیستمی_analytics_cat1"] = analytics.get("cat1") or ""
    out["سیستمی_analytics_cat2"] = analytics.get("cat2") or ""
    out["سیستمی_analytics_cat3"] = analytics.get("cat3") or ""
    out["سیستمی_city_meta_city_id"] = safe_get(detail, ["city", "city_id"]) or ""
    out["سیستمی_city_meta_name"] = safe_get(detail, ["city", "name"]) or ""

    # ویژگی‌ها تجمیع شده
    out["attributes"] = " | ".join(attrs_list)
    
    # عنوان‌های ویژگی‌ها و امکانات (ستون تجمیعی — سازگاری با نسخه قبل)
    if features_kv:
        combined_parts = [f"{k}: {v}" if v not in ("دارد", "ندارد") else f"{k} {v}" for k, v in features_kv.items()]
        out["ویژگی_ها_و_امکانات"] = " | ".join(combined_parts)
    else:
        out["ویژگی_ها_و_امکانات"] = ""

    return out


def _filter_by_advertiser_type(
    rows: List[Dict[str, Any]],
    advertiser_filter: str = "all",
    return_removed: bool = False,
) -> tuple:
    """فیلتر نوع آگهی‌دهنده: all = همه | personal = فقط شخصی (بدون مشاور/آژانس)."""
    if str(advertiser_filter or "all").strip().lower() != "personal":
        if return_removed:
            return rows, []
        return rows

    consultant_types = {
        "premium-panel", "premium_panel", "business", "consultant", "agency",
        "بنگاه", "مشاور", "آژانس",
    }
    filtered_rows: List[Dict[str, Any]] = []
    removed_ads: List[str] = []

    for row in rows:
        title = str(row.get("title", "") or row.get("عنوان", "") or "").strip()
        business_type = str(
            row.get("business_type") or row.get("فروشنده_business_type") or ""
        ).strip().lower()
        consultant_name = str(row.get("مشاور_نام") or "").strip()

        is_consultant = (
            business_type in consultant_types
            or "premium" in business_type
            or bool(consultant_name)
        )
        if is_consultant:
            if return_removed:
                removed_ads.append(title or "(بدون عنوان)")
            continue

        if business_type == "personal" or (not business_type and not consultant_name):
            filtered_rows.append(row)
        elif return_removed:
            removed_ads.append(title or "(بدون عنوان)")

    if return_removed:
        return filtered_rows, removed_ads
    return filtered_rows


def _filter_roommate_ads(rows: List[Dict[str, Any]], return_removed: bool = False) -> tuple:
    """
    فیلتر کردن آگهی‌های همخانه از لیست داده‌ها
    این آگهی‌ها برای افرادی است که دنبال هم خانه هستند و برای فایلینگ مشاوران مناسب نیستند
    
    Args:
        rows: لیست آگهی‌ها
        return_removed: اگر True باشد، لیست آگهی‌های حذف شده را هم برمی‌گرداند
    
    Returns:
        اگر return_removed=False: لیست آگهی‌های فیلتر شده
        اگر return_removed=True: (filtered_rows, removed_ads) که removed_ads لیستی از عنوان‌های حذف شده است
    """
    # بررسی اینکه آیا این آگهی‌ها مربوط به اجاره هستند
    is_rent = False
    try:
        is_rent = "rent" in str(CONFIG.get("category", "")).lower()
    except Exception:
        is_rent = False
    
    # اگر اجاره نیست، نیازی به فیلتر نیست
    if not is_rent:
        if return_removed:
            return rows, []
        return rows
    
    # الگوی regex برای تشخیص کلمات کلیدی همخانه
    # این الگو همه حالات مختلف را پوشش می‌دهد: همخانه، هم خانه، هم‌خانه، همخونه، هم خونه، هم‌خونه
    # و همچنین حالات با پسوندهای مختلف: همخانه‌ای، هم‌خانه‌ی، و غیره
    # الگو: هم (با یا بدون فاصله/خط تیره) + (خو?ن?ه یا خانه) + (اختیاری: ای/ی/‌)
    roommate_pattern = re.compile(
        r'هم[\s\-]?(?:خو?ن?ه|خانه)[\s\-]?[ای‌ی]?',
        re.IGNORECASE | re.UNICODE
    )
    
    filtered_rows = []
    removed_ads = []
    
    for row in rows:
        # دریافت عنوان و توضیحات
        title = str(row.get("title", "") or row.get("عنوان", "") or "").strip()
        description = str(row.get("description", "") or row.get("توضیحات", "") or "").strip()
        
        # ترکیب عنوان و توضیحات برای جستجو
        combined_text = f"{title} {description}"
        
        # بررسی وجود کلمات کلیدی با استفاده از regex
        if roommate_pattern.search(combined_text):
            # این آگهی همخانه است، پس حذف می‌شود
            if return_removed:
                # ذخیره عنوان برای گزارش
                ad_title = title if title else "(بدون عنوان)"
                removed_ads.append(ad_title)
            continue
        
        # اگر آگهی همخانه نبود، به لیست اضافه می‌شود
        filtered_rows.append(row)
    
    if return_removed:
        return filtered_rows, removed_ads
    
    return filtered_rows


def _apply_export_column_filter(df: Any, export_columns: Optional[List[str]]) -> Any:
    """محدود کردن ستون‌های خروجی فقط وقتی صریحاً درخواست شده (نه در استخراج عادی)."""
    if not export_columns or not len(getattr(df, "columns", [])):
        return df
    pick = [c for c in df.columns if c in export_columns]
    return df[pick] if pick else df


def export_to_excel(
    rows: List[Dict[str, Any]],
    output_path: str,
    export_columns: Optional[List[str]] = None,
) -> str:
    try:
        import pandas as pd
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"])
        import pandas as pd

    # فیلتر کردن آگهی‌های همخانه قبل از پردازش
    rows = _filter_roommate_ads(rows, return_removed=False)
    rows = normalize_mixed_export_rows(rows)

    df = pd.DataFrame(rows)
    df = df.reset_index(drop=True)
    is_rent = False
    try:
        is_rent = "rent" in str(CONFIG.get("category", "")).lower()
    except Exception:
        is_rent = False
    
    # اگر اجاره است، از منطق قدیمی استفاده می‌کنیم
    if is_rent:
        return _export_to_excel_rent(df, output_path, export_columns=export_columns)
    
    # برای فروش، از منطق جدید استفاده می‌کنیم
    return _export_to_excel_sell(df, output_path, export_columns=export_columns)


def export_to_csv(rows: List[Dict[str, Any]], output_path: str) -> str:
    """خروجی CSV از داده‌های استخراج شده - با همان پردازش Excel"""
    try:
        import pandas as pd
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas"])
        import pandas as pd
    
    # فیلتر کردن آگهی‌های همخانه قبل از پردازش
    rows = _filter_roommate_ads(rows, return_removed=False)
    rows = normalize_mixed_export_rows(rows)

    if not rows:
        raise ValueError("داده‌ای برای خروجی وجود ندارد")
    
    df = pd.DataFrame(rows)
    
    # استفاده از همان منطق پردازش Excel
    is_rent = False
    try:
        is_rent = "rent" in str(CONFIG.get("category", "")).lower()
    except Exception:
        is_rent = False
    
    # پردازش داده‌ها با همان منطق Excel
    if is_rent:
        result_df = _process_data_for_rent(df)
    else:
        result_df = _process_data_for_sell(df)
    
    # ذخیره فایل CSV با encoding UTF-8 و BOM برای سازگاری با Excel
    result_df.to_csv(output_path, index=False, encoding='utf-8-sig', sep=',')
    
    return output_path


def export_to_json(rows: List[Dict[str, Any]], output_path: str) -> str:
    """خروجی JSON از داده‌های استخراج شده - با همان پردازش Excel"""
    try:
        import pandas as pd
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas"])
        import pandas as pd
    
    # فیلتر کردن آگهی‌های همخانه قبل از پردازش
    rows = _filter_roommate_ads(rows, return_removed=False)
    rows = normalize_mixed_export_rows(rows)

    if not rows:
        raise ValueError("داده‌ای برای خروجی وجود ندارد")
    
    df = pd.DataFrame(rows)
    
    # استفاده از همان منطق پردازش Excel
    is_rent = False
    try:
        is_rent = "rent" in str(CONFIG.get("category", "")).lower()
    except Exception:
        is_rent = False
    
    # پردازش داده‌ها با همان منطق Excel
    if is_rent:
        result_df = _process_data_for_rent(df)
    else:
        result_df = _process_data_for_sell(df)
    
    # تبدیل DataFrame به لیست دیکشنری
    result_df = result_df.fillna("-")
    for col in result_df.columns:
        result_df[col] = result_df[col].apply(lambda x: "-" if (x is None or (isinstance(x, str) and str(x).strip() == "")) else x)
    
    # تبدیل به لیست دیکشنری برای JSON
    json_rows = result_df.to_dict('records')
    
    # ذخیره فایل JSON با encoding UTF-8 و فرمت خوانا
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(json_rows, f, ensure_ascii=False, indent=2)
    
    return output_path


def _is_raw_scrape_row(row: Dict[str, Any]) -> bool:
    """ردیف خام flatten هنوز به ستون‌های نهایی اکسل تبدیل نشده است."""
    if not row:
        return False
    raw_markers = (
        "attributes",
        "share_link",
        "مکان_شهر",
        "مکان_منطقه",
        "انتشار_unavailable_after",
        "مالی_قیمت_کل_عدد",
        "ویژگی_ها_و_امکانات",
    )
    if not any(row.get(key) not in (None, "", "-") for key in raw_markers):
        return False
    export_markers = ("ایجاد آگهی", "انقضا آگهی", "قیمت", "ودیعه", "اجاره ماهانه")
    has_export_values = any(
        row.get(key) not in (None, "", "-") for key in export_markers if key in row
    )
    return not has_export_values


def export_rows_to_records(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """تبدیل ردیف‌های خام استخراج به فرمت نهایی ستون‌های اکسل."""
    if not rows:
        return []
    try:
        import pandas as pd
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"])
        import pandas as pd

    df = pd.DataFrame(rows)
    is_rent = "rent" in str(CONFIG.get("category", "")).lower()
    if is_rent:
        result_df = _process_data_for_rent(df)
    else:
        result_df = _process_data_for_sell(df)

    result_df = result_df.fillna("-")
    for col in result_df.columns:
        result_df[col] = result_df[col].apply(
            lambda x: "-" if (x is None or (isinstance(x, str) and str(x).strip() == "")) else x
        )
    return result_df.to_dict("records")


def normalize_mixed_export_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """پردازش ردیف‌های خام در میان ردیف‌های از قبل export‌شده (جلوگیری از خالی شدن ستون‌ها)."""
    if not rows:
        return rows
    raw_indices = [i for i, row in enumerate(rows) if _is_raw_scrape_row(row)]
    if not raw_indices:
        return rows
    raw_rows = [rows[i] for i in raw_indices]
    processed = export_rows_to_records(raw_rows)
    normalized = list(rows)
    for idx, row in zip(raw_indices, processed):
        normalized[idx] = row
    return normalized


def merge_with_existing_excel_rows(
    new_rows: List[Dict[str, Any]],
    existing_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """ادغام ردیف‌های جدید با ردیف‌های موجود؛ ردیف‌های جدید قبل از ادغام به فرمت اکسل تبدیل می‌شوند."""
    if not existing_rows:
        return new_rows
    if not new_rows:
        return existing_rows

    existing_tokens: Set[str] = set()
    for row in existing_rows:
        token = str(row.get("token") or row.get("شناسه") or "").strip()
        if token:
            existing_tokens.add(token)

    unique_new_rows: List[Dict[str, Any]] = []
    for row in new_rows:
        token = str(row.get("token") or row.get("شناسه") or "").strip()
        if token and token not in existing_tokens:
            unique_new_rows.append(row)
            existing_tokens.add(token)

    if not unique_new_rows:
        return existing_rows

    processed_new = export_rows_to_records(unique_new_rows)
    return processed_new + existing_rows


def _process_data_for_sell(df: Any) -> Any:
    """پردازش داده‌ها برای فروش - استخراج منطق از _export_to_excel_sell"""
    import pandas as pd
    
    # تبدیل NaN به None
    df = df.where(pd.notnull(df), None)
    
    # استفاده از همان منطق _export_to_excel_sell اما بدون ذخیره Excel
    # این تابع فقط DataFrame پردازش شده را برمی‌گرداند
    # برای سادگی، از همان تابع استفاده می‌کنیم اما به جای ذخیره Excel، DataFrame را برمی‌گردانیم
    
    # ایجاد یک فایل موقت برای استفاده از منطق موجود
    import tempfile
    temp_path = tempfile.mktemp(suffix='.xlsx')
    
    try:
        _export_to_excel_sell(df, temp_path)
        # خواندن فایل Excel پردازش شده
        result_df = pd.read_excel(temp_path)
        return result_df
    finally:
        # حذف فایل موقت
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except:
            pass


def _process_data_for_rent(df: Any) -> Any:
    """پردازش داده‌ها برای اجاره - استخراج منطق از _export_to_excel_rent"""
    import pandas as pd
    
    # تبدیل NaN به None
    df = df.where(pd.notnull(df), None)
    
    # استفاده از همان منطق _export_to_excel_rent اما بدون ذخیره Excel
    # ایجاد یک فایل موقت برای استفاده از منطق موجود
    import tempfile
    temp_path = tempfile.mktemp(suffix='.xlsx')
    
    try:
        _export_to_excel_rent(df, temp_path)
        # خواندن فایل Excel پردازش شده
        result_df = pd.read_excel(temp_path)
        return result_df
    finally:
        # حذف فایل موقت
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except:
            pass


def _export_to_excel_sell(df: Any, output_path: str, export_columns: Optional[List[str]] = None) -> str:
    """خروجی اکسل برای فروش مسکونی با فرمت مشخص شده"""
    import pandas as pd
    from datetime import datetime, timedelta
    
    # تبدیل NaN به None
    df = df.where(pd.notnull(df), None)
    df = df.reset_index(drop=True)

    # استخراج و ساخت ستون‌های مورد نیاز
    result_df = pd.DataFrame(index=df.index)

    # Helper function برای دسترسی به ستون‌ها
    def get_col(col_name, default=""):
        if col_name in df.columns:
            return df[col_name]
        return pd.Series([default] * len(df), index=df.index)

    # Helper function برای ترکیب ستون اکسل و ستون اصلی
    def combine_excel_and_source(excel_col_name, source_col_name, default="-", transform_func=None):
        """ترکیب ستون اکسل و ستون اصلی: اگر ستون اکسل وجود دارد، از آن استفاده می‌کنیم.
        برای ردیف‌هایی که مقدار ندارند (NaN یا "-")، از ستون اصلی استفاده می‌کنیم."""
        if excel_col_name in df.columns:
            excel_col = get_col(excel_col_name).copy()
            # برای ردیف‌هایی که مقدار ندارند، از ستون اصلی استفاده می‌کنیم
            if source_col_name in df.columns and source_col_name != excel_col_name:
                source_col = get_col(source_col_name)
                if transform_func:
                    source_col = source_col.apply(transform_func)
                # پر کردن مقادیر خالی از ستون اصلی
                mask = excel_col.isna() | (excel_col == "") | (excel_col == "-")
                excel_col.loc[mask] = source_col.loc[mask]
            return excel_col
        else:
            # اگر ستون اکسل وجود ندارد، از ستون اصلی استفاده می‌کنیم
            if source_col_name in df.columns:
                source_col = get_col(source_col_name)
                if transform_func:
                    return source_col.apply(transform_func)
                return source_col
            return pd.Series([default] * len(df), index=df.index)

    # Helper function برای استخراج عدد از رشته متراژ
    def extract_number_from_metraj(val):
        """استخراج عدد از رشته متراژ (مثلاً "۲۴۴ متر" -> "۲۴۴")"""
        if pd.isna(val) or val == "" or val is None or val == "-":
            return "-"
        val_str = str(val).strip()
        if not val_str or val_str == "-":
            return "-"

        # حذف کلمه "متر" و فاصله‌ها
        val_str = val_str.replace("متر", "").strip()

        # استخراج اعداد (فارسی و انگلیسی)
        numbers = re.findall(r'[\d۰-۹]+', val_str)
        if numbers:
            persian_to_english = str.maketrans('۰۱۲۳۴۵۶۷۸۹', '0123456789')
            number = numbers[0].translate(persian_to_english)
            return number
        return "-"

    # 1. شناسه
    result_df["شناسه"] = combine_excel_and_source("شناسه", "token")
    
    # 2. عنوان
    result_df["عنوان"] = combine_excel_and_source("عنوان", "title")
    
    # 3. لینک
    result_df["لینک"] = combine_excel_and_source("لینک", "share_link")
    
    # 4. شماره تماس (ستون خالی)
    result_df["شماره تماس"] = pd.Series(["-"] * len(df))
    
    # 5. ایجاد آگهی
    if "ایجاد آگهی" in df.columns:
        # اگر ستون اکسل وجود دارد، از آن استفاده می‌کنیم
        excel_col = get_col("ایجاد آگهی").copy()
        # برای ردیف‌هایی که مقدار ندارند یا "-" هستند، از ستون اصلی استفاده می‌کنیم
        if "انتشار_unavailable_after" in df.columns:
            expiry_col = get_col("انتشار_unavailable_after")
            try:
                import jdatetime
            except ImportError:
                import subprocess
                subprocess.check_call([sys.executable, "-m", "pip", "install", "jdatetime"])
                import jdatetime
            
            def calc_created_from_expiry(val):
                if pd.isna(val) or val == "" or val is None:
                    return None
                try:
                    val_str = str(val).strip()
                    formats = ["%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"]
                    val_clean = val_str.split("+")[0].split("Z")[0]
                    if "T" in val_clean:
                        parts = val_clean.split("T")
                        if len(parts) == 2:
                            date_part = parts[0]
                            time_part = parts[1].split(".")[0]
                            val_clean = f"{date_part} {time_part}"
                    
                    dt = None
                    for fmt in formats:
                        try:
                            dt = datetime.strptime(val_clean.strip(), fmt)
                            break
                        except:
                            continue
                    
                    if dt:
                        created_dt = dt - timedelta(days=30)
                        jdt = jdatetime.datetime.fromgregorian(datetime=created_dt)
                        return jdt.strftime("%Y/%m/%d %H:%M:%S")
                except:
                    pass
                return None
            
            # برای ردیف‌هایی که مقدار ندارند یا "-" هستند، از ستون اصلی استفاده می‌کنیم
            calculated_col = expiry_col.apply(calc_created_from_expiry)
            excel_col = excel_col.fillna(calculated_col)
            # همچنین مقادیر "-" را با مقادیر محاسبه شده جایگزین می‌کنیم
            mask = (excel_col == "-") | (excel_col == "") | excel_col.isna()
            excel_col[mask] = calculated_col[mask]
            
            result_df["ایجاد آگهی"] = excel_col
        else:
            result_df["ایجاد آگهی"] = excel_col
    else:
        # محاسبه از زیرعنوان یا انقضا
        try:
            import jdatetime
        except ImportError:
            import subprocess
            subprocess.check_call([sys.executable, "-m", "pip", "install", "jdatetime"])
            import jdatetime
        
        def convert_to_shamsi(val):
            """تبدیل تاریخ میلادی به شمسی"""
            if pd.isna(val) or val == "" or val is None:
                return "-"
            val_str = str(val).strip()
            if not val_str or val_str == "-":
                return "-"
            try:
                # تلاش برای پارس کردن تاریخ میلادی
                formats = ["%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"]
                val_clean = val_str.split("+")[0].split("Z")[0]
                if "T" in val_clean:
                    parts = val_clean.split("T")
                    if len(parts) == 2:
                        date_part = parts[0]
                        time_part = parts[1].split(".")[0]
                        val_clean = f"{date_part} {time_part}"
                
                dt = None
                for fmt in formats:
                    try:
                        dt = datetime.strptime(val_clean.strip(), fmt)
                        break
                    except:
                        continue
                
                if dt:
                    jdt = jdatetime.datetime.fromgregorian(datetime=dt)
                    return jdt.strftime("%Y/%m/%d %H:%M:%S")
                return val_str
            except:
                return val_str
        
        # ایجاد آگهی: از انقضا 30 روز کم می‌کنیم
        expiry_col = get_col("انتشار_unavailable_after")
        created_dates = []
        if len(expiry_col) > 0 and expiry_col.notna().any():
            for val in expiry_col:
                if pd.isna(val) or val == "" or val is None:
                    created_dates.append("-")
                    continue
                try:
                    val_str = str(val).strip()
                    formats = ["%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"]
                    val_clean = val_str.split("+")[0].split("Z")[0]
                    if "T" in val_clean:
                        parts = val_clean.split("T")
                        if len(parts) == 2:
                            date_part = parts[0]
                            time_part = parts[1].split(".")[0]
                            val_clean = f"{date_part} {time_part}"
                    
                    dt = None
                    for fmt in formats:
                        try:
                            dt = datetime.strptime(val_clean.strip(), fmt)
                            break
                        except:
                            continue
                    
                    if dt:
                        # 30 روز قبل از انقضا = تاریخ ایجاد
                        created_dt = dt - timedelta(days=30)
                        jdt = jdatetime.datetime.fromgregorian(datetime=created_dt)
                        created_dates.append(jdt.strftime("%Y/%m/%d %H:%M:%S"))
                    else:
                        created_dates.append("-")
                except:
                    created_dates.append("-")
        else:
            created_dates = ["-"] * len(df)
        result_df["ایجاد آگهی"] = created_dates
    
    # 6. انقضا آگهی
    if "انقضا آگهی" in df.columns:
        # اگر ستون اکسل وجود دارد، از آن استفاده می‌کنیم
        excel_col = get_col("انقضا آگهی").copy()
        if "انتشار_unavailable_after" in df.columns:
            expiry_col = get_col("انتشار_unavailable_after")
            try:
                import jdatetime
            except ImportError:
                import subprocess
                subprocess.check_call([sys.executable, "-m", "pip", "install", "jdatetime"])
                import jdatetime
            
            def convert_to_shamsi(val):
                """تبدیل تاریخ میلادی به شمسی"""
                if pd.isna(val) or val == "" or val is None:
                    return None
                val_str = str(val).strip()
                if not val_str or val_str == "-":
                    return None
                try:
                    formats = ["%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"]
                    val_clean = val_str.split("+")[0].split("Z")[0]
                    if "T" in val_clean:
                        parts = val_clean.split("T")
                        if len(parts) == 2:
                            date_part = parts[0]
                            time_part = parts[1].split(".")[0]
                            val_clean = f"{date_part} {time_part}"
                    
                    dt = None
                    for fmt in formats:
                        try:
                            dt = datetime.strptime(val_clean.strip(), fmt)
                            break
                        except:
                            continue
                    
                    if dt:
                        jdt = jdatetime.datetime.fromgregorian(datetime=dt)
                        return jdt.strftime("%Y/%m/%d %H:%M:%S")
                    return None
                except:
                    return None
            
            # برای ردیف‌هایی که مقدار ندارند یا "-" هستند، از ستون اصلی استفاده می‌کنیم
            calculated_col = expiry_col.apply(convert_to_shamsi)
            excel_col = excel_col.fillna(calculated_col)
            # همچنین مقادیر "-" را با مقادیر محاسبه شده جایگزین می‌کنیم
            mask = (excel_col == "-") | (excel_col == "") | excel_col.isna()
            excel_col[mask] = calculated_col[mask]
            
            result_df["انقضا آگهی"] = excel_col
        else:
            result_df["انقضا آگهی"] = excel_col
    else:
        # تبدیل به شمسی
        try:
            import jdatetime
        except ImportError:
            import subprocess
            subprocess.check_call([sys.executable, "-m", "pip", "install", "jdatetime"])
            import jdatetime
        
        def convert_to_shamsi(val):
            """تبدیل تاریخ میلادی به شمسی"""
            if pd.isna(val) or val == "" or val is None:
                return "-"
            val_str = str(val).strip()
            if not val_str or val_str == "-":
                return "-"
            try:
                formats = ["%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"]
                val_clean = val_str.split("+")[0].split("Z")[0]
                if "T" in val_clean:
                    parts = val_clean.split("T")
                    if len(parts) == 2:
                        date_part = parts[0]
                        time_part = parts[1].split(".")[0]
                        val_clean = f"{date_part} {time_part}"
                
                dt = None
                for fmt in formats:
                    try:
                        dt = datetime.strptime(val_clean.strip(), fmt)
                        break
                    except:
                        continue
                
                if dt:
                    jdt = jdatetime.datetime.fromgregorian(datetime=dt)
                    return jdt.strftime("%Y/%m/%d %H:%M:%S")
                return val_str
            except:
                return val_str
        
        expiry_col = get_col("انتشار_unavailable_after")
        if len(expiry_col) > 0 and expiry_col.notna().any():
            result_df["انقضا آگهی"] = expiry_col.apply(convert_to_shamsi)
        else:
            result_df["انقضا آگهی"] = pd.Series(["-"] * len(df))
    
    # 7. دسته‌بندی   
    result_df["دسته‌بندی"] = combine_excel_and_source("دسته‌بندی", "دسته‌بندی_عنوان")
    
    # 8. شهر
    if "شهر" in df.columns and "مکان_شهر" in df.columns:
        result_df["شهر"] = combine_excel_and_source("شهر", "مکان_شهر")
    elif "شهر" in df.columns:
        result_df["شهر"] = combine_excel_and_source("شهر", "city_persian")
    else:
        result_df["شهر"] = get_col("مکان_شهر") if "مکان_شهر" in df.columns else get_col("city_persian")
    
    # 9. محله
    if "محله" in df.columns and "مکان_منطقه" in df.columns:
        result_df["محله"] = combine_excel_and_source("محله", "مکان_منطقه")
    elif "محله" in df.columns:
        result_df["محله"] = combine_excel_and_source("محله", "district_persian")
    else:
        result_df["محله"] = get_col("مکان_منطقه") if "مکان_منطقه" in df.columns else get_col("district_persian")
    
    # 10. متراژ
    if "متراژ" in df.columns:
        excel_col = get_col("متراژ").copy()
        # اعمال استخراج عدد بر روی ستون اکسل
        excel_col = excel_col.apply(extract_number_from_metraj)
        # برای ردیف‌هایی که مقدار ندارند، از attributes استخراج می‌کنیم
        if "attributes" in df.columns:
            metras = []
            for idx, raw in enumerate(df["attributes"].fillna("")):
                if pd.isna(excel_col.iloc[idx]) or excel_col.iloc[idx] == "" or excel_col.iloc[idx] == "-":
                    val = "-"
                    try:
                        parts = [p.strip() for p in str(raw).split("|") if p.strip()]
                        for p in parts:
                            if ":" in p:
                                k, v = p.split(":", 1)
                                if k.strip() == "متراژ":
                                    val = extract_number_from_metraj(v.strip())
                                    break
                    except:
                        pass
                    metras.append(val)
                else:
                    metras.append(excel_col.iloc[idx])
            result_df["متراژ"] = pd.Series(metras)
        else:
            result_df["متراژ"] = excel_col
    else:
        # اگر ستون اکسل وجود ندارد، از attributes استخراج می‌کنیم
        metras = []
        if "attributes" in df.columns:
            for raw in df["attributes"].fillna(""):
                val = "-"
                try:
                    parts = [p.strip() for p in str(raw).split("|") if p.strip()]
                    for p in parts:
                        if ":" in p:
                            k, v = p.split(":", 1)
                            if k.strip() == "متراژ":
                                val = extract_number_from_metraj(v.strip())
                                break
                except:
                    pass
                metras.append(val)
        else:
            metras = ["-"] * len(df)
        result_df["متراژ"] = pd.Series(metras)
    
    # 11. قیمت
    result_df["قیمت"] = combine_excel_and_source("قیمت", "مالی_قیمت_کل_عدد")
    
    # 12. قیمت هر متر
    result_df["قیمت هر متر"] = combine_excel_and_source("قیمت هر متر", "مالی_قیمت_هر_متر_عدد")
    
    # 13. آسانسور
    result_df["آسانسور"] = combine_excel_and_source("آسانسور", "آسانسور", "-")
    
    # 14. پارکینگ
    result_df["پارکینگ"] = combine_excel_and_source("پارکینگ", "پارکینگ", "-")
    
    # 15. انباری
    result_df["انباری"] = combine_excel_and_source("انباری", "انباری", "-")
    
    # 16. نوع آگهی دهنده
    if "نوع آگهی دهنده" in df.columns:
        excel_col = get_col("نوع آگهی دهنده").copy()
        # برای ردیف‌هایی که مقدار ندارند، از ستون اصلی استفاده می‌کنیم
        source_col_name = "فروشنده_business_type" if "فروشنده_business_type" in df.columns else "business_type"
        if source_col_name in df.columns:
            source_col = get_col(source_col_name)
            def convert_business_type(x):
                if pd.isna(x) or x == "" or x is None:
                    return "-"
                x_str = str(x).strip().lower()
                if x_str == "personal":
                    return "شخصی"
                elif x_str == "business":
                    return "بنگاه"
                elif x_str == "premium-panel":
                    return "مشاور"
                else:
                    return x if x else "-"
            converted_col = source_col.apply(convert_business_type)
            # پر کردن مقادیر خالی از ستون اصلی
            mask = excel_col.isna() | (excel_col == "") | (excel_col == "-")
            excel_col[mask] = converted_col[mask]
        result_df["نوع آگهی دهنده"] = excel_col
    else:
        business_type = get_col("فروشنده_business_type") if "فروشنده_business_type" in df.columns else get_col("business_type")
        def convert_business_type(x):
            if pd.isna(x) or x == "" or x is None:
                return "-"
            x_str = str(x).strip().lower()
            if x_str == "personal":
                return "شخصی"
            elif x_str == "business":
                return "بنگاه"
            elif x_str == "premium-panel":
                return "مشاور"
            else:
                return x if x else "-"
        result_df["نوع آگهی دهنده"] = business_type.apply(convert_business_type)

    # 16b. اطلاعات مشاور / آژانس (از API lazy)
    result_df["نام مشاور / آژانس"] = combine_excel_and_source("نام مشاور / آژانس", "مشاور_نام")
    result_df["آگهی‌های فعال مشاور"] = combine_excel_and_source("آگهی‌های فعال مشاور", "مشاور_آگهی_فعال")
    
    # 17. عرض جغرافیایی
    if "عرض جغرافیایی" in df.columns and "مکان_مختصات_lat" in df.columns:
        result_df["عرض جغرافیایی"] = combine_excel_and_source("عرض جغرافیایی", "مکان_مختصات_lat")
    elif "عرض جغرافیایی" in df.columns:
        result_df["عرض جغرافیایی"] = combine_excel_and_source("عرض جغرافیایی", "latitude")
    else:
        result_df["عرض جغرافیایی"] = get_col("مکان_مختصات_lat") if "مکان_مختصات_lat" in df.columns else get_col("latitude")
    
    # 18. طول جغرافیایی
    if "طول جغرافیایی" in df.columns and "مکان_مختصات_lng" in df.columns:
        result_df["طول جغرافیایی"] = combine_excel_and_source("طول جغرافیایی", "مکان_مختصات_lng")
    elif "طول جغرافیایی" in df.columns:
        result_df["طول جغرافیایی"] = combine_excel_and_source("طول جغرافیایی", "longitude")
    else:
        result_df["طول جغرافیایی"] = get_col("مکان_مختصات_lng") if "مکان_مختصات_lng" in df.columns else get_col("longitude")

    result_df["عرض جغرافیایی"] = result_df["عرض جغرافیایی"].apply(_normalize_geo_coord)
    result_df["طول جغرافیایی"] = result_df["طول جغرافیایی"].apply(_normalize_geo_coord)
    
    # 19. نوع مختصات
    if "نوع مختصات" in df.columns:
        excel_col = get_col("نوع مختصات").copy()
        if "نقشه_نوع" in df.columns:
            map_type = get_col("نقشه_نوع")
            converted_col = map_type.apply(lambda x: "دقیق" if x == "EXACT" else ("تقریبی" if x == "FUZZY" else (x if x else "-")))
            mask = excel_col.isna() | (excel_col == "") | (excel_col == "-")
            excel_col[mask] = converted_col[mask]
        result_df["نوع مختصات"] = excel_col
    else:
        map_type = get_col("نقشه_نوع")
        result_df["نوع مختصات"] = map_type.apply(lambda x: "دقیق" if x == "EXACT" else ("تقریبی" if x == "FUZZY" else (x if x else "-")))
    
    # 20. در شعاع (متر)
    result_df["در شعاع (متر)"] = combine_excel_and_source("در شعاع (متر)", "نقشه_شعاع_محدوده_متر")
    
    # 21. تعداد تصاویر
    result_df["تعداد تصاویر"] = combine_excel_and_source("تعداد تصاویر", "رسانه_تعداد_تصویر")
    
    # 22. اصالت تصویر
    result_df["اصالت تصویر"] = combine_excel_and_source("اصالت تصویر", "رسانه_اصالت_تصویر_برای_همین_ملک")
    
    # 23. ویدیو دارد؟
    if "ویدیو دارد؟" in df.columns:
        excel_col = get_col("ویدیو دارد؟").copy()
        if "رسانه_ویدیو" in df.columns:
            video_col = get_col("رسانه_ویدیو")
            converted_col = video_col.apply(lambda x: "دارد" if x and str(x).strip() else "ندارد")
            mask = excel_col.isna() | (excel_col == "") | (excel_col == "-")
            excel_col[mask] = converted_col[mask]
        result_df["ویدیو دارد؟"] = excel_col
    else:
        video_col = get_col("رسانه_ویدیو")
        result_df["ویدیو دارد؟"] = video_col.apply(lambda x: "دارد" if x and str(x).strip() else "ندارد")
    
    # 24. ویژگی‌ها و امکانات (تجمیعی)
    result_df["ویژگی‌ها و امکانات"] = combine_excel_and_source("ویژگی‌ها و امکانات", "ویژگی_ها_و_امکانات")
    _append_feature_detail_columns(result_df, df, combine_excel_and_source)
    
    # 25. توضیحات
    if "توضیحات" in df.columns:
        result_df["توضیحات"] = combine_excel_and_source("توضیحات", "description")
    else:
        result_df["توضیحات"] = get_col("توضیحات") if "توضیحات" in df.columns else get_col("description")
    
    # 26. تصویر نقشه
    result_df["تصویر نقشه"] = combine_excel_and_source("تصویر نقشه", "نقشه_تصویر")
    
    # ترتیب ستون‌ها — ویژگی‌های تفکیک‌شده بعد از امکانات پایه
    sell_column_order = [
        "شناسه", "عنوان", "لینک", "شماره تماس", "ایجاد آگهی", "انقضا آگهی",
        "دسته‌بندی", "شهر", "محله", "متراژ", "قیمت", "قیمت هر متر",
        "آسانسور", "پارکینگ", "انباری",
        *FEATURE_EXCEL_COLUMNS,
        "نوع آگهی دهنده", "نام مشاور / آژانس", "آگهی‌های فعال مشاور",
        "عرض جغرافیایی", "طول جغرافیایی", "نوع مختصات",
        "در شعاع (متر)", "تعداد تصاویر", "اصالت تصویر", "ویدیو دارد؟",
        "ویژگی‌ها و امکانات", "توضیحات", "تصویر نقشه",
    ]
    ordered = [c for c in sell_column_order if c in result_df.columns]
    rest = [c for c in result_df.columns if c not in ordered]
    result_df = result_df[ordered + rest]
    
    # تبدیل مقادیر خالی به "-"
    result_df = result_df.fillna("-")
    for col in result_df.columns:
        result_df[col] = result_df[col].apply(lambda x: "-" if (x is None or (isinstance(x, str) and str(x).strip() == "")) else x)
    
    result_df = _apply_export_column_filter(result_df, export_columns)
    
    # ذخیره فایل
    result_df.to_excel(output_path, index=False)
    
    # فرمت‌بندی حرفه‌ای فایل
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # تعریف استایل‌ها
        # استفاده از فونت استاندارد برای سازگاری بیشتر
        try:
            header_font = Font(name='B Nazanin', size=11, bold=True, color='FFFFFF')
        except:
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # آبی حرفه‌ای
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        numeric_alignment = Alignment(horizontal='center', vertical='center')
        text_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        text_alignment_no_wrap = Alignment(horizontal='right', vertical='center', wrap_text=False)
        
        # تعریف ستون‌های عددی (که باید وسط‌چین باشند)
        integer_numeric_columns = [
            "شناسه", "متراژ", "قیمت", "قیمت هر متر",
            "در شعاع (متر)", "تعداد تصاویر"
        ]
        
        # تعریف ستون‌های boolean-like و تاریخ (که باید وسط‌چین باشند)
        center_columns = [
            "ویدیو دارد؟", "آسانسور", "پارکینگ", "انباری", "نوع مختصات",
            "ایجاد آگهی", "انقضا آگهی", "شماره تماس"
        ]
        
        # ستون‌هایی که نباید wrap_text داشته باشند (برای نمایش یک خطی)
        no_wrap_columns = [
            "توضیحات", "ویژگی‌ها و امکانات", "تصویر نقشه"
        ]
        
        # فرمت‌بندی هدر (ردیف اول)
        for col_idx in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col_idx)
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_text = header_cell.value
            if header_text == "قیمت":
                header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            elif header_text in no_wrap_columns:
                header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            else:
                header_cell.alignment = header_alignment
        
        # فرمت‌بندی داده‌ها
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                header_text = ws.cell(row=1, column=col_idx).value
                
                # اگر ستون عددی است، وسط‌چین
                if _apply_geo_coord_format(cell, header_text, numeric_alignment):
                    pass
                elif header_text in integer_numeric_columns and not _match_geo_header(header_text):
                    # برای ستون قیمت، wrap_text را غیرفعال می‌کنیم
                    if header_text == "قیمت":
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    else:
                        cell.alignment = numeric_alignment
                    # اگر مقدار عددی است، فرمت عددی اعمال می‌کنیم
                    try:
                        if cell.value and str(cell.value).strip() != "-" and str(cell.value).strip() != "":
                            val = str(cell.value).replace(',', '').replace(' ', '')
                            if val.replace('.', '').replace('-', '').isdigit():
                                cell.number_format = '#,##0'
                    except:
                        pass
                elif header_text in center_columns:
                    # ستون‌های boolean-like وسط‌چین
                    cell.alignment = numeric_alignment
                elif header_text in no_wrap_columns:
                    # ستون‌های متنی بدون wrap (یک خطی)
                    cell.alignment = text_alignment_no_wrap
                else:
                    # ستون‌های متنی راست‌چین
                    cell.alignment = text_alignment
        
        # تنظیم ارتفاع ردیف‌ها
        # ارتفاع استاندارد برای هدر
        ws.row_dimensions[1].height = 18
        # ارتفاع استاندارد برای ردیف‌های داده
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 15
        
        # تنظیم عرض ستون‌ها
        features_col_idx = None
        price_col_idx = None
        for col_idx, col_cells in enumerate(ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=ws.max_row), start=1):
            max_len = 0
            header_text = ws.cell(row=1, column=col_idx).value
            # بررسی اینکه آیا این ستون "ویژگی‌ها و امکانات" است
            if header_text == "ویژگی‌ها و امکانات":
                features_col_idx = col_idx
            # بررسی اینکه آیا این ستون "قیمت" است
            if header_text == "قیمت":
                price_col_idx = col_idx
            
            for cell in col_cells:
                value = cell.value
                if value is None:
                    continue
                text = str(value)
                longest_line = max((len(line) for line in text.splitlines() if line is not None), default=len(text))
                if longest_line > max_len:
                    max_len = longest_line
            # حداقل عرض 12 و حداکثر 50
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 50)
        
        # دو برابر کردن عرض ستون "ویژگی‌ها و امکانات"
        if features_col_idx:
            current_width = ws.column_dimensions[get_column_letter(features_col_idx)].width
            ws.column_dimensions[get_column_letter(features_col_idx)].width = current_width * 2
        
        # افزایش عرض ستون "قیمت" (1.5 برابر)
        if price_col_idx:
            current_width = ws.column_dimensions[get_column_letter(price_col_idx)].width
            ws.column_dimensions[get_column_letter(price_col_idx)].width = current_width * 1.5
        
        # اضافه کردن border به همه سلول‌ها
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        # border برای هدر
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=1, column=col_idx).border = thin_border
        # پیدا کردن ستون "نوع آگهی دهنده"
        business_type_col_idx = None
        for col_idx in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col_idx)
            if header_cell.value == "نوع آگهی دهنده":
                business_type_col_idx = col_idx
                break
        
        # رنگ سبز کمرنگ برای سطرهای شخصی
        light_green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        
        # border برای داده‌ها
        for row_idx in range(2, ws.max_row + 1):
            # بررسی نوع آگهی دهنده
            is_personal = False
            if business_type_col_idx:
                business_type_cell = ws.cell(row=row_idx, column=business_type_col_idx)
                if business_type_cell.value == "شخصی":
                    is_personal = True
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                
                # اگر شخصی است، پس‌زمینه سبز کمرنگ
                if is_personal:
                    cell.fill = light_green_fill
                else:
                    # رنگ پس‌زمینه متناوب برای خوانایی بهتر (فقط برای غیرشخصی)
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
        
        # فریز کردن ردیف اول (هدر)
        ws.freeze_panes = 'A2'

        _fix_worksheet_geo_coordinates(ws, numeric_alignment)
        
        wb.save(output_path)
    except Exception as e:
        # در صورت خطا، حداقل عرض ستون‌ها را تنظیم می‌کنیم
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            wb = load_workbook(output_path)
            ws = wb.active
            for col_idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            _fix_worksheet_geo_coordinates(ws)
            wb.save(output_path)
        except:
            pass
    
    return output_path


def _export_to_excel_rent(df: Any, output_path: str, export_columns: Optional[List[str]] = None) -> str:
    """خروجی اکسل برای اجاره مسکونی با فرمت مشخص شده"""
    import pandas as pd
    from datetime import datetime, timedelta
    
    # تبدیل NaN به None
    df = df.where(pd.notnull(df), None)
    df = df.reset_index(drop=True)
    
    # استخراج و ساخت ستون‌های مورد نیاز
    result_df = pd.DataFrame(index=df.index)
    
    # Helper function برای دسترسی به ستون‌ها
    def get_col(col_name, default=""):
        if col_name in df.columns:
            return df[col_name]
        return pd.Series([default] * len(df), index=df.index)
    
    # Helper function برای ترکیب ستون اکسل و ستون اصلی
    def combine_excel_and_source(excel_col_name, source_col_name, default="-", transform_func=None):
        """ترکیب ستون اکسل و ستون اصلی: اگر ستون اکسل وجود دارد، از آن استفاده می‌کنیم.
        برای ردیف‌هایی که مقدار ندارند (NaN یا "-")، از ستون اصلی استفاده می‌کنیم."""
        if excel_col_name in df.columns:
            excel_col = get_col(excel_col_name).copy()
            # برای ردیف‌هایی که مقدار ندارند، از ستون اصلی استفاده می‌کنیم
            if source_col_name in df.columns and source_col_name != excel_col_name:
                source_col = get_col(source_col_name)
                if transform_func:
                    source_col = source_col.apply(transform_func)
                # پر کردن مقادیر خالی از ستون اصلی
                mask = excel_col.isna() | (excel_col == "") | (excel_col == "-")
                excel_col.loc[mask] = source_col.loc[mask]
            return excel_col
        else:
            # اگر ستون اکسل وجود ندارد، از ستون اصلی استفاده می‌کنیم
            if source_col_name in df.columns:
                source_col = get_col(source_col_name)
                if transform_func:
                    return source_col.apply(transform_func)
                return source_col
            return pd.Series([default] * len(df), index=df.index)
    
    # Helper function برای استخراج عدد از رشته متراژ
    def extract_number_from_metraj(val):
        """استخراج عدد از رشته متراژ (مثلاً "۲۴۴ متر" -> "۲۴۴")"""
        if pd.isna(val) or val == "" or val is None or val == "-":
            return "-"
        val_str = str(val).strip()
        if not val_str or val_str == "-":
            return "-"
        
        # حذف کلمه "متر" و فاصله‌ها
        val_str = val_str.replace("متر", "").strip()
        
        # استخراج اعداد (فارسی و انگلیسی)
        numbers = re.findall(r'[\d۰-۹]+', val_str)
        if numbers:
            # تبدیل اعداد فارسی به انگلیسی
            persian_to_english = str.maketrans('۰۱۲۳۴۵۶۷۸۹', '0123456789')
            number = numbers[0].translate(persian_to_english)
            return number
        return "-"
    
    # Helper function برای بررسی اینکه آیا مقدار عددی خالص است
    def is_pure_numeric(val):
        """بررسی اینکه آیا مقدار عددی خالص است (int یا float)"""
        if pd.isna(val) or val == "" or val is None:
            return False
        # بررسی اینکه آیا مقدار از نوع عددی است
        if isinstance(val, (int, float)):
            return True
        # بررسی اینکه آیا رشته فقط شامل اعداد است (بدون کاراکترهای اضافی)
        val_str = str(val).strip()
        if not val_str or val_str == "-":
            return False
        # حذف فاصله‌ها و بررسی اینکه آیا فقط اعداد است
        val_clean = val_str.replace(" ", "").replace(",", "").replace("،", "")
        # بررسی اینکه آیا فقط اعداد فارسی یا انگلیسی دارد
        if val_clean.isdigit():
            return True
        # بررسی اعداد فارسی
        persian_digits = '۰۱۲۳۴۵۶۷۸۹'
        if all(c in persian_digits or c in '0123456789' for c in val_clean):
            # بررسی اینکه آیا فقط اعداد است (بدون کاراکترهای دیگر)
            if all(c in persian_digits + '0123456789' for c in val_clean):
                return True
        return False
    
    # Helper function برای استخراج عدد از رشته قیمت (فقط برای مقادیر رشته‌ای)
    def extract_number_from_price(val):
        """استخراج عدد از رشته قیمت (مثلاً "۱،۲۰۰،۰۰۰،۰۰۰ تومان" -> "1200000000")
        اگر مقدار عددی خالص باشد، بدون تغییر برمی‌گرداند."""
        if pd.isna(val) or val == "" or val is None or val == "-":
            return "-"
        
        # اگر مقدار عددی خالص است، آن را به رشته تبدیل می‌کنیم و برمی‌گردانیم
        if is_pure_numeric(val):
            if isinstance(val, (int, float)):
                return str(int(val))
            val_str = str(val).strip()
            # تبدیل اعداد فارسی به انگلیسی
            persian_to_english = str.maketrans('۰۱۲۳۴۵۶۷۸۹', '0123456789')
            val_str = val_str.translate(persian_to_english)
            # حذف فاصله‌ها و کاماها
            val_clean = val_str.replace(" ", "").replace(",", "").replace("،", "")
            return val_clean
        
        val_str = str(val).strip()
        if not val_str or val_str == "-":
            return "-"
        
        # حذف کلمات و کاراکترهای غیر عددی
        val_str = val_str.replace("تومان", "").replace("ریال", "").replace(",", "").replace("،", "").strip()
        
        # استخراج اعداد (فارسی و انگلیسی)
        numbers = re.findall(r'[\d۰-۹]+', val_str)
        if numbers:
            # تبدیل اعداد فارسی به انگلیسی
            persian_to_english = str.maketrans('۰۱۲۳۴۵۶۷۸۹', '0123456789')
            number = "".join([n.translate(persian_to_english) for n in numbers])
            return number
        return "-"
    
    # Helper function برای تبدیل قابل تبدیل/غیر قابل تبدیل
    def convert_convertible(val):
        """تبدیل مقدار قابل تبدیل/غیر قابل تبدیل"""
        if pd.isna(val) or val == "" or val is None or val == "-":
            return "-"
        val_str = str(val).strip()
        # اول بررسی دقیق مقادیر اصلی
        if val_str == "قابل تبدیل":
            return "قابل تبدیل"
        elif val_str == "غیر قابل تبدیل" or val_str == "غیرقابل تبدیل":
            return "غیر قابل تبدیل"
        # سپس بررسی وجود کلمات کلیدی در رشته
        # اول "غیر قابل تبدیل" را بررسی می‌کنیم (اولویت با غیر قابل تبدیل)
        if "غیر قابل تبدیل" in val_str or "غیرقابل تبدیل" in val_str:
            return "غیر قابل تبدیل"
        # سپس "قابل تبدیل" را بررسی می‌کنیم
        elif "قابل تبدیل" in val_str:
            return "قابل تبدیل"
        # بررسی انگلیسی
        val_lower = val_str.lower()
        if "non-convertible" in val_lower or "nonconvertible" in val_lower:
            return "غیر قابل تبدیل"
        elif "convertible" in val_lower:
            return "قابل تبدیل"
        # در غیر این صورت، مقدار اصلی را برگردان
        return val_str
    
    # Helper function برای استخراج مقدار از attributes
    def extract_from_attributes(attr_key, default="-"):
        """استخراج مقدار از attributes بر اساس کلید"""
        if "attributes" not in df.columns:
            return pd.Series([default] * len(df))
        results = []
        for raw in df["attributes"].fillna(""):
            val = default
            try:
                parts = [p.strip() for p in str(raw).split("|") if p.strip()]
                for p in parts:
                    if ":" in p:
                        k, v = p.split(":", 1)
                        if k.strip() == attr_key:
                            val = v.strip()
                            break
            except:
                pass
            results.append(val)
        return pd.Series(results)
    
    # 1. شناسه
    result_df["شناسه"] = combine_excel_and_source("شناسه", "token")
    
    # 2. عنوان
    result_df["عنوان"] = combine_excel_and_source("عنوان", "title")
    
    # 3. شهر
    if "شهر" in df.columns and "مکان_شهر" in df.columns:
        result_df["شهر"] = combine_excel_and_source("شهر", "مکان_شهر")
    elif "شهر" in df.columns:
        result_df["شهر"] = combine_excel_and_source("شهر", "city_persian")
    else:
        result_df["شهر"] = get_col("مکان_شهر") if "مکان_شهر" in df.columns else get_col("city_persian")
    
    # 4. محله
    if "محله" in df.columns and "مکان_منطقه" in df.columns:
        result_df["محله"] = combine_excel_and_source("محله", "مکان_منطقه")
    elif "محله" in df.columns:
        result_df["محله"] = combine_excel_and_source("محله", "district_persian")
    else:
        result_df["محله"] = get_col("مکان_منطقه") if "مکان_منطقه" in df.columns else get_col("district_persian")
    
    # 5. آدرس
    result_df["آدرس"] = combine_excel_and_source("آدرس", "مکان_آدرس", "-")
    
    # 6. متراژ
    if "متراژ" in df.columns:
        excel_col = get_col("متراژ").copy()
        excel_col = excel_col.apply(extract_number_from_metraj)
        if "attributes" in df.columns:
            metras = []
            for idx, raw in enumerate(df["attributes"].fillna("")):
                if pd.isna(excel_col.iloc[idx]) or excel_col.iloc[idx] == "" or excel_col.iloc[idx] == "-":
                    val = extract_from_attributes("متراژ", "-").iloc[idx]
                    if val != "-":
                        val = extract_number_from_metraj(val)
                    metras.append(val)
                else:
                    metras.append(excel_col.iloc[idx])
            result_df["متراژ"] = pd.Series(metras)
        else:
            result_df["متراژ"] = excel_col
    else:
        result_df["متراژ"] = extract_from_attributes("متراژ", "-").apply(extract_number_from_metraj)
    
    # 7. ساخت
    if "ساخت" in df.columns:
        excel_col = get_col("ساخت").copy()
        # فقط برای ردیف‌هایی که مقدار ندارند و attributes دارند، از attributes استفاده می‌کنیم
        if "attributes" in df.columns:
            mask = excel_col.isna() | (excel_col == "") | (excel_col == "-")
            if mask.any():
                # فقط برای ردیف‌هایی که attributes دارند و مقدار ندارند
                for idx in df.index[mask]:
                    if pd.notna(df.loc[idx, "attributes"]) and str(df.loc[idx, "attributes"]).strip():
                        val = "-"
                        try:
                            raw = str(df.loc[idx, "attributes"]).strip()
                            parts = [p.strip() for p in raw.split("|") if p.strip()]
                            for p in parts:
                                if ":" in p:
                                    k, v = p.split(":", 1)
                                    if k.strip() == "ساخت":
                                        val = v.strip()
                                        break
                        except:
                            pass
                        if val != "-":
                            excel_col.iloc[idx] = val
        result_df["ساخت"] = excel_col
    else:
        result_df["ساخت"] = extract_from_attributes("ساخت", "-")
    
    # 8. اتاق
    if "اتاق" in df.columns:
        excel_col = get_col("اتاق").copy()
        # فقط برای ردیف‌هایی که مقدار ندارند و attributes دارند، از attributes استفاده می‌کنیم
        if "attributes" in df.columns:
            mask = excel_col.isna() | (excel_col == "") | (excel_col == "-")
            if mask.any():
                # فقط برای ردیف‌هایی که attributes دارند و مقدار ندارند
                for idx in df.index[mask]:
                    if pd.notna(df.loc[idx, "attributes"]) and str(df.loc[idx, "attributes"]).strip():
                        val = "-"
                        try:
                            raw = str(df.loc[idx, "attributes"]).strip()
                            parts = [p.strip() for p in raw.split("|") if p.strip()]
                            for p in parts:
                                if ":" in p:
                                    k, v = p.split(":", 1)
                                    if k.strip() == "اتاق":
                                        val = v.strip()
                                        break
                        except:
                            pass
                        if val != "-":
                            excel_col.iloc[idx] = val
        result_df["اتاق"] = excel_col
    else:
        result_df["اتاق"] = extract_from_attributes("اتاق", "-")
    
    # 9. طبقه
    if "طبقه" in df.columns:
        excel_col = get_col("طبقه").copy()
        # فقط برای ردیف‌هایی که مقدار ندارند و attributes دارند، از attributes استفاده می‌کنیم
        if "attributes" in df.columns:
            mask = excel_col.isna() | (excel_col == "") | (excel_col == "-")
            if mask.any():
                # فقط برای ردیف‌هایی که attributes دارند و مقدار ندارند
                for idx in df.index[mask]:
                    if pd.notna(df.loc[idx, "attributes"]) and str(df.loc[idx, "attributes"]).strip():
                        val = "-"
                        try:
                            raw = str(df.loc[idx, "attributes"]).strip()
                            parts = [p.strip() for p in raw.split("|") if p.strip()]
                            for p in parts:
                                if ":" in p:
                                    k, v = p.split(":", 1)
                                    if k.strip() == "طبقه":
                                        val = v.strip()
                                        break
                        except:
                            pass
                        if val != "-":
                            excel_col.iloc[idx] = val
        result_df["طبقه"] = excel_col
    else:
        result_df["طبقه"] = extract_from_attributes("طبقه", "-")
    
    # 10. ودیعه
    if "ودیعه" in df.columns:
        excel_col = get_col("ودیعه").copy()
        # فقط روی مقادیر خالی یا "-" پردازش انجام می‌شود
        # مقادیر موجود که عددی خالص هستند، بدون تغییر نگه داشته می‌شوند
        mask_empty = excel_col.isna() | (excel_col == "") | (excel_col == "-")
        # برای مقادیر غیر خالی که رشته‌ای با فرمت خاص هستند، پردازش انجام می‌شود
        mask_to_process = ~mask_empty
        if mask_to_process.any():
            for idx in df.index[mask_to_process]:
                val = excel_col.iloc[idx]
                # فقط اگر مقدار عددی خالص نیست، پردازش می‌شود
                if not is_pure_numeric(val):
                    processed_val = extract_number_from_price(val)
                    excel_col.iloc[idx] = processed_val
        
        # برای مقادیر خالی، از ستون منبع استفاده می‌کنیم
        if "اجاره_ودیعه_عدد" in df.columns:
            source_col = get_col("اجاره_ودیعه_عدد")
            excel_col[mask_empty] = source_col[mask_empty]
        result_df["ودیعه"] = excel_col
    else:
        result_df["ودیعه"] = get_col("اجاره_ودیعه_عدد", "-")
    
    # 11. اجاره ماهانه
    if "اجاره ماهانه" in df.columns:
        excel_col = get_col("اجاره ماهانه").copy()
        # فقط روی مقادیر خالی یا "-" پردازش انجام می‌شود
        # مقادیر موجود که عددی خالص هستند، بدون تغییر نگه داشته می‌شوند
        mask_empty = excel_col.isna() | (excel_col == "") | (excel_col == "-")
        # برای مقادیر غیر خالی که رشته‌ای با فرمت خاص هستند، پردازش انجام می‌شود
        mask_to_process = ~mask_empty
        if mask_to_process.any():
            for idx in df.index[mask_to_process]:
                val = excel_col.iloc[idx]
                # فقط اگر مقدار عددی خالص نیست، پردازش می‌شود
                if not is_pure_numeric(val):
                    processed_val = extract_number_from_price(val)
                    excel_col.iloc[idx] = processed_val
        
        # برای مقادیر خالی، از ستون منبع استفاده می‌کنیم
        if "اجاره_ماهانه_عدد" in df.columns:
            source_col = get_col("اجاره_ماهانه_عدد")
            excel_col[mask_empty] = source_col[mask_empty]
        result_df["اجاره ماهانه"] = excel_col
    else:
        result_df["اجاره ماهانه"] = get_col("اجاره_ماهانه_عدد", "-")
    
    # 12. ودیعه و اجاره (قابل تبدیل/غیر قابل تبدیل)
    if "ودیعه و اجاره" in df.columns:
        excel_col = get_col("ودیعه و اجاره").copy()
        excel_col = excel_col.apply(convert_convertible)
        if "اجاره_قابل_تبدیل" in df.columns:
            source_col = get_col("اجاره_قابل_تبدیل").apply(convert_convertible)
            mask = excel_col.isna() | (excel_col == "") | (excel_col == "-")
            excel_col[mask] = source_col[mask]
        result_df["ودیعه و اجاره"] = excel_col
    else:
        result_df["ودیعه و اجاره"] = get_col("اجاره_قابل_تبدیل", "-").apply(convert_convertible)
    
    # 13. آسانسور
    result_df["آسانسور"] = combine_excel_and_source("آسانسور", "آسانسور", "-")
    
    # 14. پارکینگ
    result_df["پارکینگ"] = combine_excel_and_source("پارکینگ", "پارکینگ", "-")
    
    # 15. انباری
    result_df["انباری"] = combine_excel_and_source("انباری", "انباری", "-")
    
    # 16. ویژگی‌ها و امکانات (تجمیعی)
    result_df["ویژگی‌ها و امکانات"] = combine_excel_and_source("ویژگی‌ها و امکانات", "ویژگی_ها_و_امکانات")
    _append_feature_detail_columns(result_df, df, combine_excel_and_source)
    
    # 17. توضیحات
    if "توضیحات" in df.columns:
        result_df["توضیحات"] = combine_excel_and_source("توضیحات", "description")
    else:
        result_df["توضیحات"] = get_col("توضیحات") if "توضیحات" in df.columns else get_col("description")
    
    # 18. عرض جغرافیایی
    if "عرض جغرافیایی" in df.columns and "مکان_مختصات_lat" in df.columns:
        result_df["عرض جغرافیایی"] = combine_excel_and_source("عرض جغرافیایی", "مکان_مختصات_lat")
    elif "عرض جغرافیایی" in df.columns:
        result_df["عرض جغرافیایی"] = combine_excel_and_source("عرض جغرافیایی", "latitude")
    else:
        result_df["عرض جغرافیایی"] = get_col("مکان_مختصات_lat") if "مکان_مختصات_lat" in df.columns else get_col("latitude")
    
    # 19. طول جغرافیایی
    if "طول جغرافیایی" in df.columns and "مکان_مختصات_lng" in df.columns:
        result_df["طول جغرافیایی"] = combine_excel_and_source("طول جغرافیایی", "مکان_مختصات_lng")
    elif "طول جغرافیایی" in df.columns:
        result_df["طول جغرافیایی"] = combine_excel_and_source("طول جغرافیایی", "longitude")
    else:
        result_df["طول جغرافیایی"] = get_col("مکان_مختصات_lng") if "مکان_مختصات_lng" in df.columns else get_col("longitude")

    result_df["عرض جغرافیایی"] = result_df["عرض جغرافیایی"].apply(_normalize_geo_coord)
    result_df["طول جغرافیایی"] = result_df["طول جغرافیایی"].apply(_normalize_geo_coord)
    
    # 20. در شعاع (متر)
    result_df["در شعاع (متر)"] = combine_excel_and_source("در شعاع (متر)", "نقشه_شعاع_محدوده_متر")
    
    # 21. نوع مختصات
    if "نوع مختصات" in df.columns:
        excel_col = get_col("نوع مختصات").copy()
        if "نقشه_نوع" in df.columns:
            map_type = get_col("نقشه_نوع")
            converted_col = map_type.apply(lambda x: "دقیق" if x == "EXACT" else ("تقریبی" if x == "FUZZY" else (x if x else "-")))
            mask = excel_col.isna() | (excel_col == "") | (excel_col == "-")
            excel_col[mask] = converted_col[mask]
        result_df["نوع مختصات"] = excel_col
    else:
        map_type = get_col("نقشه_نوع")
        result_df["نوع مختصات"] = map_type.apply(lambda x: "دقیق" if x == "EXACT" else ("تقریبی" if x == "FUZZY" else (x if x else "-")))
    
    # 22. تعداد تصاویر
    result_df["تعداد تصاویر"] = combine_excel_and_source("تعداد تصاویر", "رسانه_تعداد_تصویر")
    
    # 23. ویدیو دارد؟
    if "ویدیو دارد؟" in df.columns:
        excel_col = get_col("ویدیو دارد؟").copy()
        if "رسانه_ویدیو" in df.columns:
            video_col = get_col("رسانه_ویدیو")
            converted_col = video_col.apply(lambda x: "دارد" if x and str(x).strip() else "ندارد")
            mask = excel_col.isna() | (excel_col == "") | (excel_col == "-")
            excel_col[mask] = converted_col[mask]
        result_df["ویدیو دارد؟"] = excel_col
    else:
        video_col = get_col("رسانه_ویدیو")
        result_df["ویدیو دارد؟"] = video_col.apply(lambda x: "دارد" if x and str(x).strip() else "ندارد")
    
    # 24. ایجاد آگهی (محاسبه از انقضا - 30 روز)
    try:
        import jdatetime
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "jdatetime"])
        import jdatetime
    
    if "ایجاد آگهی" in df.columns:
        # اگر ستون اکسل وجود دارد، از آن استفاده می‌کنیم
        excel_col = get_col("ایجاد آگهی").copy()
        # برای ردیف‌هایی که مقدار ندارند یا "-" هستند، از ستون اصلی استفاده می‌کنیم
        if "انتشار_unavailable_after" in df.columns:
            expiry_col = get_col("انتشار_unavailable_after")
            
            def calc_created_from_expiry(val):
                if pd.isna(val) or val == "" or val is None:
                    return None
                try:
                    val_str = str(val).strip()
                    formats = ["%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"]
                    val_clean = val_str.split("+")[0].split("Z")[0]
                    if "T" in val_clean:
                        parts = val_clean.split("T")
                        if len(parts) == 2:
                            date_part = parts[0]
                            time_part = parts[1].split(".")[0]
                            val_clean = f"{date_part} {time_part}"
                    
                    dt = None
                    for fmt in formats:
                        try:
                            dt = datetime.strptime(val_clean.strip(), fmt)
                            break
                        except:
                            continue
                    
                    if dt:
                        created_dt = dt - timedelta(days=30)
                        jdt = jdatetime.datetime.fromgregorian(datetime=created_dt)
                        return jdt.strftime("%Y/%m/%d %H:%M:%S")
                except:
                    pass
                return None
            
            # برای ردیف‌هایی که مقدار ندارند یا "-" هستند، از ستون اصلی استفاده می‌کنیم
            calculated_col = expiry_col.apply(calc_created_from_expiry)
            excel_col = excel_col.fillna(calculated_col)
            # همچنین مقادیر "-" را با مقادیر محاسبه شده جایگزین می‌کنیم
            mask = (excel_col == "-") | (excel_col == "") | excel_col.isna()
            excel_col[mask] = calculated_col[mask]
            
            result_df["ایجاد آگهی"] = excel_col
        else:
            result_df["ایجاد آگهی"] = excel_col
    else:
        # محاسبه از انقضا
        def convert_to_shamsi(val):
            """تبدیل تاریخ میلادی به شمسی"""
            if pd.isna(val) or val == "" or val is None:
                return "-"
            val_str = str(val).strip()
            if not val_str or val_str == "-":
                return "-"
            try:
                formats = ["%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"]
                val_clean = val_str.split("+")[0].split("Z")[0]
                if "T" in val_clean:
                    parts = val_clean.split("T")
                    if len(parts) == 2:
                        date_part = parts[0]
                        time_part = parts[1].split(".")[0]
                        val_clean = f"{date_part} {time_part}"
                
                dt = None
                for fmt in formats:
                    try:
                        dt = datetime.strptime(val_clean.strip(), fmt)
                        break
                    except:
                        continue
                
                if dt:
                    jdt = jdatetime.datetime.fromgregorian(datetime=dt)
                    return jdt.strftime("%Y/%m/%d %H:%M:%S")
                return val_str
            except:
                return val_str
        
        # ایجاد آگهی: از انقضا 30 روز کم می‌کنیم
        expiry_col = get_col("انتشار_unavailable_after")
        created_dates = []
        if len(expiry_col) > 0 and expiry_col.notna().any():
            for val in expiry_col:
                if pd.isna(val) or val == "" or val is None:
                    created_dates.append("-")
                    continue
                try:
                    val_str = str(val).strip()
                    formats = ["%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"]
                    val_clean = val_str.split("+")[0].split("Z")[0]
                    if "T" in val_clean:
                        parts = val_clean.split("T")
                        if len(parts) == 2:
                            date_part = parts[0]
                            time_part = parts[1].split(".")[0]
                            val_clean = f"{date_part} {time_part}"
                    
                    dt = None
                    for fmt in formats:
                        try:
                            dt = datetime.strptime(val_clean.strip(), fmt)
                            break
                        except:
                            continue
                    
                    if dt:
                        # 30 روز قبل از انقضا = تاریخ ایجاد
                        created_dt = dt - timedelta(days=30)
                        jdt = jdatetime.datetime.fromgregorian(datetime=created_dt)
                        created_dates.append(jdt.strftime("%Y/%m/%d %H:%M:%S"))
                    else:
                        created_dates.append("-")
                except:
                    created_dates.append("-")
        else:
            created_dates = ["-"] * len(df)
        result_df["ایجاد آگهی"] = created_dates
    
    # 25. انقضا آگهی
    if "انقضا آگهی" in df.columns:
        # اگر ستون اکسل وجود دارد، از آن استفاده می‌کنیم
        excel_col = get_col("انقضا آگهی").copy()
        if "انتشار_unavailable_after" in df.columns:
            expiry_col = get_col("انتشار_unavailable_after")
            try:
                import jdatetime
            except ImportError:
                import subprocess
                subprocess.check_call([sys.executable, "-m", "pip", "install", "jdatetime"])
                import jdatetime
            
            def convert_to_shamsi(val):
                """تبدیل تاریخ میلادی به شمسی"""
                if pd.isna(val) or val == "" or val is None:
                    return None
                val_str = str(val).strip()
                if not val_str or val_str == "-":
                    return None
                try:
                    formats = ["%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"]
                    val_clean = val_str.split("+")[0].split("Z")[0]
                    if "T" in val_clean:
                        parts = val_clean.split("T")
                        if len(parts) == 2:
                            date_part = parts[0]
                            time_part = parts[1].split(".")[0]
                            val_clean = f"{date_part} {time_part}"
                    
                    dt = None
                    for fmt in formats:
                        try:
                            dt = datetime.strptime(val_clean.strip(), fmt)
                            break
                        except:
                            continue
                    
                    if dt:
                        jdt = jdatetime.datetime.fromgregorian(datetime=dt)
                        return jdt.strftime("%Y/%m/%d %H:%M:%S")
                    return None
                except:
                    return None
            
            # برای ردیف‌هایی که مقدار ندارند یا "-" هستند، از ستون اصلی استفاده می‌کنیم
            calculated_col = expiry_col.apply(convert_to_shamsi)
            excel_col = excel_col.fillna(calculated_col)
            # همچنین مقادیر "-" را با مقادیر محاسبه شده جایگزین می‌کنیم
            mask = (excel_col == "-") | (excel_col == "") | excel_col.isna()
            excel_col[mask] = calculated_col[mask]
            
            result_df["انقضا آگهی"] = excel_col
        else:
            result_df["انقضا آگهی"] = excel_col
    else:
        # تبدیل به شمسی
        try:
            import jdatetime
        except ImportError:
            import subprocess
            subprocess.check_call([sys.executable, "-m", "pip", "install", "jdatetime"])
            import jdatetime
        
        def convert_to_shamsi(val):
            """تبدیل تاریخ میلادی به شمسی"""
            if pd.isna(val) or val == "" or val is None:
                return "-"
            val_str = str(val).strip()
            if not val_str or val_str == "-":
                return "-"
            try:
                formats = ["%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"]
                val_clean = val_str.split("+")[0].split("Z")[0]
                if "T" in val_clean:
                    parts = val_clean.split("T")
                    if len(parts) == 2:
                        date_part = parts[0]
                        time_part = parts[1].split(".")[0]
                        val_clean = f"{date_part} {time_part}"
                
                dt = None
                for fmt in formats:
                    try:
                        dt = datetime.strptime(val_clean.strip(), fmt)
                        break
                    except:
                        continue
                
                if dt:
                    jdt = jdatetime.datetime.fromgregorian(datetime=dt)
                    return jdt.strftime("%Y/%m/%d %H:%M:%S")
                return val_str
            except:
                return val_str
        
        expiry_col = get_col("انتشار_unavailable_after")
        if len(expiry_col) > 0 and expiry_col.notna().any():
            result_df["انقضا آگهی"] = expiry_col.apply(convert_to_shamsi)
        else:
            result_df["انقضا آگهی"] = pd.Series(["-"] * len(df))
    
    # 26. نوع آگهی دهنده
    if "نوع آگهی دهنده" in df.columns:
        excel_col = get_col("نوع آگهی دهنده").copy()
        # برای ردیف‌هایی که مقدار ندارند، از ستون اصلی استفاده می‌کنیم
        source_col_name = "فروشنده_business_type" if "فروشنده_business_type" in df.columns else "business_type"
        if source_col_name in df.columns:
            source_col = get_col(source_col_name)
            def convert_business_type(x):
                if pd.isna(x) or x == "" or x is None:
                    return "-"
                x_str = str(x).strip().lower()
                if x_str == "personal":
                    return "شخصی"
                elif x_str == "business":
                    return "بنگاه"
                elif x_str == "premium-panel":
                    return "مشاور"
                else:
                    return x if x else "-"
            converted_col = source_col.apply(convert_business_type)
            # پر کردن مقادیر خالی از ستون اصلی
            mask = excel_col.isna() | (excel_col == "") | (excel_col == "-")
            excel_col[mask] = converted_col[mask]
        result_df["نوع آگهی دهنده"] = excel_col
    else:
        business_type = get_col("فروشنده_business_type") if "فروشنده_business_type" in df.columns else get_col("business_type")
        def convert_business_type(x):
            if pd.isna(x) or x == "" or x is None:
                return "-"
            x_str = str(x).strip().lower()
            if x_str == "personal":
                return "شخصی"
            elif x_str == "business":
                return "بنگاه"
            elif x_str == "premium-panel":
                return "مشاور"
            else:
                return x if x else "-"
        result_df["نوع آگهی دهنده"] = business_type.apply(convert_business_type)

    # 26b. اطلاعات مشاور / آژانس (از API lazy)
    result_df["نام مشاور / آژانس"] = combine_excel_and_source("نام مشاور / آژانس", "مشاور_نام")
    result_df["آگهی‌های فعال مشاور"] = combine_excel_and_source("آگهی‌های فعال مشاور", "مشاور_آگهی_فعال")
    
    # 26. لینک
    result_df["لینک"] = get_col("لینک")
    
    # 27. شماره تماس (ستون خالی)
    result_df["شماره تماس"] = pd.Series(["-"] * len(df))
    
    # تبدیل مقادیر خالی به "-"
    result_df = result_df.fillna("-")
    for col in result_df.columns:
        result_df[col] = result_df[col].apply(lambda x: "-" if (x is None or (isinstance(x, str) and str(x).strip() == "")) else x)
    
    # تنظیم ترتیب ستون‌ها
    column_order = [
        "شناسه", "عنوان", "لینک", "نوع آگهی دهنده",
        "نام مشاور / آژانس", "آگهی‌های فعال مشاور",
        "شهر", "محله", "آدرس", "متراژ", "ساخت", "اتاق", "طبقه",
        "آسانسور", "پارکینگ", "انباری",
        *FEATURE_EXCEL_COLUMNS,
        "ودیعه", "اجاره ماهانه", "ودیعه و اجاره",
        "ایجاد آگهی", "انقضا آگهی", "شماره تماس",
        "تعداد تصاویر", "ویدیو دارد؟",
        "عرض جغرافیایی", "طول جغرافیایی", "در شعاع (متر)", "نوع مختصات",
        "ویژگی‌ها و امکانات", "توضیحات",
    ]
    
    # فقط ستون‌هایی که در result_df وجود دارند را اضافه می‌کنیم
    ordered_columns = [col for col in column_order if col in result_df.columns]
    # ستون‌های اضافی که در column_order نیستند را به انتها اضافه می‌کنیم
    remaining_columns = [col for col in result_df.columns if col not in ordered_columns]
    result_df = result_df[ordered_columns + remaining_columns]
    
    result_df = _apply_export_column_filter(result_df, export_columns)
    
    # ذخیره فایل
    result_df.to_excel(output_path, index=False)
    
    # فرمت‌بندی حرفه‌ای فایل (مشابه فروش)
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # تعریف استایل‌ها
        try:
            header_font = Font(name='B Nazanin', size=11, bold=True, color='FFFFFF')
        except:
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        numeric_alignment = Alignment(horizontal='center', vertical='center')
        text_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        text_alignment_no_wrap = Alignment(horizontal='right', vertical='center', wrap_text=False)
        
        # تعریف ستون‌های عددی
        integer_numeric_columns = [
            "شناسه", "متراژ", "ودیعه", "اجاره ماهانه",
            "در شعاع (متر)", "تعداد تصاویر", "ساخت", "اتاق", "طبقه"
        ]
        
        # تعریف ستون‌های boolean-like و تاریخ
        center_columns = [
            "ویدیو دارد؟", "آسانسور", "پارکینگ", "انباری", "نوع مختصات",
            "ایجاد آگهی", "انقضا آگهی", "شماره تماس", "ودیعه و اجاره"
        ]
        
        # ستون‌هایی که نباید wrap_text داشته باشند (برای نمایش یک خطی)
        no_wrap_columns = [
            "توضیحات", "ویژگی‌ها و امکانات"
        ]
        
        # فرمت‌بندی هدر
        for col_idx in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col_idx)
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_text = header_cell.value
            if header_text in no_wrap_columns:
                header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            else:
                header_cell.alignment = header_alignment
        
        # فرمت‌بندی داده‌ها
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                header_text = ws.cell(row=1, column=col_idx).value
                
                if _apply_geo_coord_format(cell, header_text, numeric_alignment):
                    pass
                elif header_text in integer_numeric_columns and not _match_geo_header(header_text):
                    cell.alignment = numeric_alignment
                    try:
                        if cell.value and str(cell.value).strip() != "-" and str(cell.value).strip() != "":
                            val = str(cell.value).replace(',', '').replace(' ', '')
                            if val.replace('.', '').replace('-', '').isdigit():
                                cell.number_format = '#,##0'
                    except:
                        pass
                elif header_text in center_columns:
                    cell.alignment = numeric_alignment
                elif header_text in no_wrap_columns:
                    # ستون‌های متنی بدون wrap (یک خطی) - اعمال صریح wrap_text=False
                    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)
                else:
                    cell.alignment = text_alignment
        
        # تنظیم ارتفاع ردیف‌ها
        ws.row_dimensions[1].height = 18
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 15
        
        # تنظیم عرض ستون‌ها
        features_col_idx = None
        description_col_idx = None
        for col_idx, col_cells in enumerate(ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=ws.max_row), start=1):
            max_len = 0
            header_text = ws.cell(row=1, column=col_idx).value
            if header_text == "ویژگی‌ها و امکانات":
                features_col_idx = col_idx
            elif header_text == "توضیحات":
                description_col_idx = col_idx
            
            for cell in col_cells:
                value = cell.value
                if value is None:
                    continue
                text = str(value)
                # برای ستون‌های no_wrap، کل طول متن را در نظر می‌گیریم (نه فقط longest_line)
                if header_text in no_wrap_columns:
                    text_len = len(text)
                    if text_len > max_len:
                        max_len = text_len
                else:
                    longest_line = max((len(line) for line in text.splitlines() if line is not None), default=len(text))
                    if longest_line > max_len:
                        max_len = longest_line
            
            # برای ستون‌های no_wrap، عرض بیشتری می‌دهیم و محدودیت 50 را حذف می‌کنیم
            if header_text in no_wrap_columns:
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 30), 100)
            else:
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 50)
        
        # دو برابر کردن عرض ستون "ویژگی‌ها و امکانات"
        if features_col_idx:
            current_width = ws.column_dimensions[get_column_letter(features_col_idx)].width
            ws.column_dimensions[get_column_letter(features_col_idx)].width = current_width * 2
        
        # اعمال مجدد wrap_text=False برای ستون‌های no_wrap (بعد از تنظیم عرض ستون‌ها)
        for col_idx in range(1, ws.max_column + 1):
            header_text = ws.cell(row=1, column=col_idx).value
            if header_text in no_wrap_columns:
                # اعمال wrap_text=False برای هدر
                header_cell = ws.cell(row=1, column=col_idx)
                header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                # اعمال wrap_text=False برای همه داده‌ها
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)
        
        # اضافه کردن border
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=1, column=col_idx).border = thin_border
        
        # پیدا کردن ستون "نوع آگهی دهنده"
        business_type_col_idx = None
        for col_idx in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col_idx)
            if header_cell.value == "نوع آگهی دهنده":
                business_type_col_idx = col_idx
                break
        
        # رنگ سبز کمرنگ برای سطرهای شخصی
        light_green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        
        # border و رنگ پس‌زمینه برای داده‌ها
        for row_idx in range(2, ws.max_row + 1):
            is_personal = False
            if business_type_col_idx:
                business_type_cell = ws.cell(row=row_idx, column=business_type_col_idx)
                if business_type_cell.value == "شخصی":
                    is_personal = True
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                
                if is_personal:
                    cell.fill = light_green_fill
                else:
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
                
                header_text = ws.cell(row=1, column=col_idx).value
                
                if _apply_geo_coord_format(cell, header_text, numeric_alignment):
                    pass
                elif header_text in integer_numeric_columns and not _match_geo_header(header_text):
                    cell.alignment = numeric_alignment
                    try:
                        if cell.value and str(cell.value).strip() != "-" and str(cell.value).strip() != "":
                            val = str(cell.value).replace(',', '').replace(' ', '')
                            if val.replace('.', '').replace('-', '').isdigit():
                                cell.number_format = '#,##0'
                    except:
                        pass
                elif header_text in center_columns:
                    cell.alignment = numeric_alignment
                else:
                    cell.alignment = text_alignment
        
        # فریز کردن ردیف اول
        ws.freeze_panes = 'A2'

        _fix_worksheet_geo_coordinates(ws, numeric_alignment)
        
        wb.save(output_path)
    except Exception as e:
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            wb = load_workbook(output_path)
            ws = wb.active
            for col_idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            _fix_worksheet_geo_coordinates(ws)
            wb.save(output_path)
        except:
            pass
    
    return output_path


def main() -> None:
    print("در حال واکشی لیست آگهی‌ها...")
    tokens = fetch_post_tokens()
    print(f"تعداد توکن‌های یافت‌شده: {len(tokens)}")

    rows: List[Dict[str, Any]] = []
    if tokens:
        session = get_http_session()
        from concurrent.futures import ThreadPoolExecutor, as_completed
        max_workers = max(1, int(CONFIG.get("max_workers", 5)))
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(fetch_post_detail, t, session): t for t in tokens}
            for idx, fut in enumerate(as_completed(futures), start=1):
                token = futures[fut]
                try:
                    detail = fut.result()
                except Exception:
                    detail = None
                if not detail:
                    continue
                row = flatten_post_detail(detail)
                rows.append(row)
                if idx % 10 == 0:
                    print(f"پردازش {idx}/{len(tokens)} ...")

    if not rows:
        print("داده‌ای برای خروجی وجود ندارد.")
        return

    # تعیین نام فایل خروجی (برای فروش و اجاره): اسلاگ منطقه در صورت وجود، در غیر این صورت اسلاگ شهر + زمان
    output_path = CONFIG["excel_output"]
    def make_slug(text: str) -> str:
        s = str(text or "").strip().lower()
        s = s.replace(" ", "-")
        s = re.sub(r"[^a-z0-9\-_]+", "-", s)
        s = re.sub(r"-+", "-", s).strip("-")
        return s or "unknown"

    rep = rows[0] if rows else {}
    # اولویت: اسلاگ انتخاب‌شده از UI → اسلاگ‌های موجود در داده
    pref_district = str(CONFIG.get("selected_district_slug") or "").strip()
    pref_city = str(CONFIG.get("selected_city_slug") or "").strip()
    district_slug = make_slug(pref_district or (rep.get("normalized_district") or ""))
    city_slug = make_slug(pref_city or (rep.get("normalized_city") or ""))
    area_part = district_slug or city_slug or "unknown"
    ts = time.strftime("%Y%m%d_%H%M%S")
    output_path = f"{area_part}_{ts}.xlsx"

    print("در حال تولید فایل اکسل...")
    final_path = export_to_excel(rows, output_path)
    print(f"✓ فایل اکسل با موفقیت ایجاد شد: {final_path}")


if __name__ == "__main__":
    main()


